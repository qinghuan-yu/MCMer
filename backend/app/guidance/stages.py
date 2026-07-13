"""Business stages for the high-trust guidance pipeline."""

from __future__ import annotations

import json
import csv
import hashlib
import re
from pathlib import Path
from typing import Protocol

from app.guidance.contracts import (
    ProblemSpec,
    VerificationReport,
)
from app.guidance.model_ir import ApprovedModelIR, ModelIR, ModelIRReview
from app.guidance.model_ir_compiler import ExecutionPlan, ModelIRCompiler
from app.guidance.model_ir_semantics import ModelIRSemanticValidator
from app.guidance.model_validators import build_default_model_validator_registry
from app.guidance.operator_execution import OperatorExecutionManifest, OperatorGraphExecutor
from app.artifacts.guidance_audit import build_guidance_audit_report
from app.guidance.renderer import render_verified_guidance


class ProblemDecomposer(Protocol):
    async def decompose(
        self,
        *,
        question: str,
        data_files: list[str],
        task: dict,
    ) -> ProblemSpec: ...


class Modeler(Protocol):
    async def model(
        self,
        *,
        problem_spec: ProblemSpec,
        data_profile: dict,
        task: dict,
        work_dir: Path,
    ) -> ModelIR: ...

    async def revise(
        self,
        *,
        problem_spec: ProblemSpec,
        previous_model_ir: ModelIR,
        review: ModelIRReview,
        data_profile: dict,
        task: dict,
        work_dir: Path,
    ) -> ModelIR: ...


class ModelReviewer(Protocol):
    async def review(
        self,
        *,
        problem_spec: ProblemSpec,
        model_ir: ModelIR,
        task: dict,
        work_dir: Path,
    ) -> ModelIRReview: ...


class DecompositionStage:
    def __init__(self, *, decomposer: ProblemDecomposer) -> None:
        self.decomposer = decomposer

    async def run(
        self,
        *,
        task_id: str,
        task: dict,
        input_path: Path | None,
        output_path: Path,
    ) -> Path:
        root = output_path.parent
        question = str(task.get("source_question") or task.get("question") or "").strip()
        data_files = _list_data_files(root)
        problem_spec = await self.decomposer.decompose(
            question=question,
            data_files=data_files,
            task=task,
        )
        problem_spec = _reconcile_source_requirements(
            problem_spec,
            question=question,
            available_files=data_files,
        )
        _write_model(output_path, problem_spec)
        return output_path


class ModelingStage:
    def __init__(self, *, modeler: Modeler) -> None:
        self.modeler = modeler

    async def run(
        self,
        *,
        task_id: str,
        task: dict,
        input_path: Path | None,
        output_path: Path,
    ) -> Path:
        source_path = _required_input(input_path)
        data_profile = _build_data_profile(output_path.parent)
        (output_path.parent / "data_profile.json").write_text(
            json.dumps(data_profile, ensure_ascii=False, indent=2),
            encoding="utf-8",
        )
        if source_path.name == "problem_spec.json":
            problem_spec = ProblemSpec.model_validate_json(source_path.read_text(encoding="utf-8"))
            model_ir = await self.modeler.model(
                problem_spec=problem_spec,
                data_profile=data_profile,
                task=task,
                work_dir=output_path.parent,
            )
        else:
            review = ModelIRReview.model_validate_json(source_path.read_text(encoding="utf-8"))
            problem_spec = ProblemSpec.model_validate_json(
                (output_path.parent / review.problem_spec_ref).read_text(encoding="utf-8")
            )
            previous_model_ir = ModelIR.model_validate_json(
                (output_path.parent / review.model_ir_ref).read_text(encoding="utf-8")
            )
            model_ir = await self.modeler.revise(
                problem_spec=problem_spec,
                previous_model_ir=previous_model_ir,
                review=review,
                data_profile=data_profile,
                task=task,
                work_dir=output_path.parent,
            )
        _validate_model_ir_coverage(problem_spec, model_ir)
        diagnostics = ModelIRSemanticValidator().validate(model_ir)
        if diagnostics:
            details = "; ".join(
                f"{item.code} at {item.path}: {item.message}" for item in diagnostics
            )
            raise ValueError(f"Model IR semantic validation failed: {details}")
        _write_model(output_path, model_ir)
        return output_path


class ModelReviewStage:
    def __init__(self, *, reviewer: ModelReviewer) -> None:
        self.reviewer = reviewer

    async def run(
        self,
        *,
        task_id: str,
        task: dict,
        input_path: Path | None,
        output_path: Path,
    ) -> Path:
        source_path = _required_input(input_path)
        model_ir = ModelIR.model_validate_json(source_path.read_text(encoding="utf-8"))
        problem_spec = ProblemSpec.model_validate_json(
            (output_path.parent / "problem_spec.json").read_text(encoding="utf-8")
        )
        review = await self.reviewer.review(
            problem_spec=problem_spec,
            model_ir=model_ir,
            task=task,
            work_dir=output_path.parent,
        )
        review = review.model_copy(
            update={
                "problem_spec_ref": "problem_spec.json",
                "model_ir_ref": source_path.name,
            }
        )
        _write_model(output_path.parent / "model_review.json", review)
        approved = ApprovedModelIR(
            review_status=review.status,
            model_id=model_ir.model_id,
            model_ir_ref=source_path.name,
            approved_model_ir=model_ir,
        )
        _write_model(output_path, approved)
        return output_path


class CalculationStage:
    """Compile approved Model IR and execute only registered operators."""

    def __init__(
        self,
        *,
        compiler: ModelIRCompiler,
        executor: OperatorGraphExecutor,
    ) -> None:
        self.compiler = compiler
        self.executor = executor

    async def run(
        self,
        *,
        task_id: str,
        task: dict,
        input_path: Path | None,
        output_path: Path,
    ) -> Path:
        source_path = _required_input(input_path)
        approved = ApprovedModelIR.model_validate_json(
            source_path.read_text(encoding="utf-8")
        )
        if approved.review_status != "approved":
            raise ValueError("CalculationStage requires an approved Model IR")
        work_dir = output_path.parent
        compilation = self.compiler.compile_to_file(
            approved.approved_model_ir,
            output_path=work_dir / "execution_plan.json",
            model_ir_ref=f"{source_path.name}#approved_model_ir",
        )
        _write_model(work_dir / "compilation_result.json", compilation)
        if compilation.execution_plan is None:
            self.executor.write_blocked(
                compilation_result=compilation,
                model_ir=approved.approved_model_ir,
                work_dir=work_dir,
            )
            return output_path
        self.executor.execute(
            execution_plan=compilation.execution_plan,
            model_ir=approved.approved_model_ir,
            work_dir=work_dir,
        )
        return output_path


class ResultVerificationStage:
    """Verify the operator execution graph without inferring model adequacy."""

    async def run(
        self,
        *,
        task_id: str,
        task: dict,
        input_path: Path | None,
        output_path: Path,
    ) -> Path:
        manifest = OperatorExecutionManifest.model_validate_json(
            _required_input(input_path).read_text(encoding="utf-8")
        )
        root = output_path.parent
        approved = ApprovedModelIR.model_validate_json(
            (root / "approved_model_ir.json").read_text(encoding="utf-8")
        )
        execution_plan = ExecutionPlan.model_validate_json(
            (root / manifest.execution_plan_ref).read_text(encoding="utf-8")
        )
        parameters = _read_json(root / "parameter_registry.json")
        results = _read_json(root / "result_registry.json")
        figure_registry = _read_json(root / "figure_registry.json")

        execution_issues: list[str] = []
        if manifest.status != "passed":
            execution_issues.append(f"operator execution status is {manifest.status}")
        missing = [
            path for path in manifest.generated_files if not (root / path).is_file()
        ]
        if missing:
            execution_issues.append(f"missing generated files: {', '.join(missing)}")
        plan_hash = hashlib.sha256(
            execution_plan.model_dump_json().encode("utf-8")
        ).hexdigest()
        if plan_hash != manifest.execution_plan_sha256:
            execution_issues.append("execution plan hash does not match manifest")
        expected_node_ids = [node.node_id for node in execution_plan.graph.nodes]
        if manifest.executed_node_ids != expected_node_ids:
            execution_issues.append("executed node IDs do not match execution plan order")

        evidence_issues, figures = _operator_evidence_issues(
            root=root,
            executed_node_ids=set(manifest.executed_node_ids),
            parameter_registry=parameters,
            result_registry=results,
            figure_registry=figure_registry,
        )
        model_adequacy = _assess_model_adequacy(approved, results)
        if execution_issues or evidence_issues:
            model_adequacy["status"] = "blocked"
            model_adequacy["reasons"].append(
                "model adequacy is blocked until execution and evidence layers pass"
            )
        counterevidence_records = _build_counterevidence_records(
            approved,
            model_adequacy,
        )

        report = VerificationReport(
            model_id=manifest.model_id,
            execution_verified={
                "status": "passed" if not execution_issues else "failed",
                "reasons": execution_issues or ["execution plan hash and node coverage verified"],
                "evidence_refs": [manifest.execution_plan_ref, input_path.name],
            },
            evidence_supported={
                "status": (
                    "blocked" if execution_issues else "passed" if not evidence_issues else "failed"
                ),
                "reasons": evidence_issues or ["registered outputs trace to executed operator nodes"],
                "evidence_refs": [
                    "parameter_registry.json",
                    "result_registry.json",
                    "figure_registry.json",
                ],
            },
            model_adequate=model_adequacy,
            counterevidence_records=counterevidence_records,
            input_checks=[{
                "status": "passed" if not missing else "failed",
                "checked_files": manifest.generated_files,
                "missing_files": missing,
            }],
            blocking_issues=[*execution_issues, *evidence_issues],
            artifact_refs={
                "approved_model": "approved_model_ir.json",
                "execution_manifest": input_path.name,
                "execution_plan": manifest.execution_plan_ref,
                "parameter_registry": "parameter_registry.json",
                "result_registry": "result_registry.json",
                "figure_registry": "figure_registry.json",
                "figures": figures,
            },
        )
        _write_model(output_path, report)
        return output_path


def _assess_model_adequacy(
    approved: ApprovedModelIR,
    result_registry: dict,
) -> dict:
    registry = build_default_model_validator_registry()
    verified_results = result_registry.get("verified_results")
    if not isinstance(verified_results, list):
        verified_results = []

    candidate_roles: set[str] = set()
    checks: dict[str, str] = {}
    reasons: list[str] = []
    evidence_refs: list[str] = []
    family_statuses: list[str] = []

    for subproblem in approved.approved_model_ir.subproblems:
        if subproblem.execution_status != "solvable":
            continue
        roles = {candidate.role for candidate in subproblem.candidate_models}
        candidate_roles.update(roles)
        subproblem_results = [
            result
            for result in verified_results
            if isinstance(result, dict)
            and result.get("subproblem_id") == subproblem.subproblem_id
        ]
        evidence = [
            result["value"]
            for result in subproblem_results
            if isinstance(result.get("value"), dict)
        ]
        for result in subproblem_results:
            result_id = str(result.get("id") or "")
            reference = f"result_registry.json#{result_id}"
            if result_id and reference not in evidence_refs:
                evidence_refs.append(reference)

        for family in subproblem.validation.model_families:
            assessment = registry.assess(
                family=family,
                candidate_roles=roles,
                evidence=evidence,
            )
            family_statuses.append(assessment.status)
            prefix = f"{subproblem.subproblem_id}.{family}"
            for name, check in assessment.checks.items():
                checks[f"{prefix}.{name}"] = check.status
            if assessment.reasons:
                reasons.extend(
                    f"{prefix}: {reason}" for reason in assessment.reasons
                )

    if family_statuses and all(status == "passed" for status in family_statuses):
        status = "passed"
        reasons.append("all declared model-family validators passed")
    elif "failed" in family_statuses:
        status = "failed"
    else:
        status = "not_assessed"
        if not reasons:
            reasons.append("no assessable model-family evidence was registered")

    return {
        "status": status,
        "candidate_roles": sorted(candidate_roles) or ["alternative"],
        "checks": checks,
        "reasons": reasons,
        "evidence_refs": evidence_refs or ["approved_model_ir.json"],
    }


def _build_counterevidence_records(
    approved: ApprovedModelIR,
    model_adequacy: dict,
) -> list[dict]:
    checks = model_adequacy.get("checks")
    if not isinstance(checks, dict):
        checks = {}
    report_evidence_refs = model_adequacy.get("evidence_refs")
    if not isinstance(report_evidence_refs, list):
        report_evidence_refs = []

    records: list[dict] = []
    for subproblem in approved.approved_model_ir.subproblems:
        if subproblem.execution_status != "solvable":
            continue
        prefix = f"{subproblem.subproblem_id}."
        subproblem_checks = {
            name: status
            for name, status in checks.items()
            if name.startswith(prefix)
        }
        passed_refs = [
            name for name, status in subproblem_checks.items() if status == "passed"
        ]
        failed_refs = [
            name for name, status in subproblem_checks.items() if status == "failed"
        ]
        baseline_comparison_passed = any(
            name.endswith(".baseline_comparison")
            for name in passed_refs
        )

        for candidate in subproblem.candidate_models:
            if candidate.role == "recommended":
                continue
            if (
                candidate.role == "baseline"
                and model_adequacy.get("status") == "passed"
                and baseline_comparison_passed
            ):
                status = "excluded"
                reason = (
                    "the recommended route passed the registered baseline comparison "
                    "and every declared model-family validator"
                )
                check_refs = passed_refs
                evidence_refs = list(report_evidence_refs)
            elif failed_refs:
                status = "not_excluded"
                reason = (
                    "the available validation evidence does not justify excluding "
                    "this route"
                )
                check_refs = failed_refs
                evidence_refs = list(report_evidence_refs)
            else:
                status = "not_assessed"
                reason = "candidate-specific exclusion evidence was not produced"
                check_refs = []
                evidence_refs = []
            records.append({
                "subproblem_id": subproblem.subproblem_id,
                "challenged_candidate_id": candidate.id,
                "status": status,
                "exclusion_reason": reason,
                "check_refs": check_refs,
                "evidence_refs": evidence_refs,
                "applicability_boundary": candidate.applicability_boundary,
            })
    return records


class GuidanceStage:
    """Render guidance only from a verification report and its artifact references."""

    async def run(
        self,
        *,
        task_id: str,
        task: dict,
        input_path: Path | None,
        output_path: Path,
    ) -> Path:
        source_path = _required_input(input_path)
        root = output_path.parent
        if source_path.name == "approved_model_ir.json":
            report, approved, parameters, results = _build_blocked_guidance_artifacts(
                root,
                source_path,
            )
        else:
            report = VerificationReport.model_validate_json(
                source_path.read_text(encoding="utf-8")
            )
            refs = report.artifact_refs
            approved = _read_json(root / str(refs["approved_model"]))
            parameters = _read_json_if_exists(root / str(refs["parameter_registry"]))
            results = _read_json_if_exists(root / str(refs["result_registry"]))
            execution_manifest = _read_json_if_exists(
                root / str(refs.get("execution_manifest", "execution_manifest.json"))
            )
            calculation_failure = _read_json_if_exists(root / "calculation_failure.json")
            report = _enforce_verified_guidance_evidence(
                report,
                approved_model=approved,
                parameter_registry=parameters,
                result_registry=results,
            )
        if source_path.name == "approved_model_ir.json":
            execution_manifest = {}
            calculation_failure = {}
        guidance = render_verified_guidance(
            approved_model=approved,
            verification_report=report.model_dump(mode="json"),
            parameter_registry=parameters,
            result_registry=results,
            execution_manifest=execution_manifest,
            calculation_failure=calculation_failure,
        )
        context = {
            "artifact_type": "guidance",
            "required_sections": _required_guidance_sections(guidance),
        }
        audit = build_guidance_audit_report(
            guidance_markdown=guidance,
            guidance_context=context,
            result_registry=results,
            parameter_registry=parameters,
            figure_registry={
                "figures": report.artifact_refs.get("figures", []),
            },
            artifact_root=root,
        )
        output_path.write_text(guidance, encoding="utf-8")
        (root / "res.md").write_text(guidance, encoding="utf-8")
        (root / "guidance_context.json").write_text(
            json.dumps(context, ensure_ascii=False, indent=2), encoding="utf-8"
        )
        (root / "guidance_audit_report.json").write_text(
            json.dumps(audit, ensure_ascii=False, indent=2), encoding="utf-8"
        )
        return output_path


def _required_guidance_sections(guidance: str) -> list[str]:
    if (
        guidance.startswith("# 计算链路阻断诊断")
        or guidance.startswith("# 建模链路阻断诊断")
        or guidance.startswith("# 结果复核阻断诊断")
    ):
        return [
            "可信度摘要",
            "阻断位置",
            "阻断原因",
            "复核检查概览" if guidance.startswith("# 结果复核阻断诊断") else "",
            "下一步修复动作",
            "可复现附件说明",
        ]
    return [
        "可信度摘要",
        "问题理解",
        "数据与来源",
        "参数与来源表",
        "模型选择理由",
        "建模路线与读者决策",
        "建模路线取舍",
        "可辨识性分析",
        "结论状态登记",
        "分问题建模步骤",
        "必要计算结果",
        "最终答案与应填表格",
        "复核与稳健性",
        "阻断项与待确认项",
        "论文转化建议",
        "可复现附件说明",
    ]


def _required_input(input_path: Path | None) -> Path:
    if input_path is None:
        raise ValueError("stage requires the previous saved contract")
    return input_path


def _validate_model_ir_coverage(problem_spec: ProblemSpec, model_ir: ModelIR) -> None:
    expected = [str(item.get("id") or "").strip() for item in problem_spec.subproblems]
    modeled = [item.subproblem_id.strip() for item in model_ir.subproblems]
    invalid_ids = not all(expected) or not all(modeled)
    duplicate_ids = len(expected) != len(set(expected)) or len(modeled) != len(set(modeled))
    if invalid_ids or duplicate_ids or set(expected) != set(modeled):
        raise ValueError(
            f"subproblem coverage mismatch: expected={expected}; modeled={modeled}"
        )


def _list_data_files(root: Path) -> list[str]:
    data_dir = root / "data"
    if not data_dir.is_dir():
        return []
    return sorted(path.relative_to(root).as_posix() for path in data_dir.rglob("*") if path.is_file())


def _reconcile_source_requirements(
    problem_spec: ProblemSpec,
    *,
    question: str,
    available_files: list[str],
) -> ProblemSpec:
    references = list(
        dict.fromkeys(
            re.sub(r"\s+", "", match.group(0))
            for match in re.finditer(r"(?i)(?:附录\s*\d+|appendix\s*[a-z0-9]+)", question)
        )
    )
    requirements = []
    missing_references: list[str] = []
    for reference in references:
        normalized_reference = re.sub(r"\s+", "", reference).casefold()
        matched_files = [
            path
            for path in available_files
            if normalized_reference in re.sub(r"\s+", "", Path(path).stem).casefold()
        ]
        embedded_heading = bool(
            re.search(
                rf"(?im)^\s*{re.escape(reference)}\s*(?:[:：]|$)",
                question,
            )
        )
        status = "available" if matched_files or embedded_heading else "missing"
        if status == "missing":
            missing_references.append(reference)
        requirements.append(
            {
                "reference": reference,
                "status": status,
                "matched_files": matched_files,
                "note": (
                    "Matched an uploaded file or an embedded appendix heading."
                    if status == "available"
                    else "Referenced by the problem statement but absent from uploaded files."
                ),
            }
        )

    locked_facts = [
        fact
        for fact in problem_spec.locked_facts
        if not any(reference in fact for reference in missing_references)
    ]
    data_sources = [
        _remove_missing_source_claims(source, missing_references)
        for source in problem_spec.data_sources
    ]
    uncertainties = list(problem_spec.uncertainties)
    for reference in missing_references:
        message = f"{reference}在题面中被引用但未上传；依赖该来源的子问题必须标记 blocked。"
        if message not in uncertainties:
            uncertainties.append(message)
    return ProblemSpec.model_validate(
        {
            **problem_spec.model_dump(mode="json"),
            "data_sources": data_sources,
            "source_requirements": requirements,
            "locked_facts": locked_facts,
            "uncertainties": uncertainties,
        }
    )


def _remove_missing_source_claims(source: dict, missing_references: list[str]) -> dict:
    cleaned: dict = {}
    for key, value in source.items():
        if isinstance(value, str):
            if not any(reference in value for reference in missing_references):
                cleaned[key] = value
        elif isinstance(value, list):
            cleaned[key] = [
                item
                for item in value
                if not isinstance(item, str)
                or not any(reference in item for reference in missing_references)
            ]
        else:
            cleaned[key] = value
    return cleaned


def _build_data_profile(root: Path) -> dict:
    files: list[dict] = []
    for path in (root / "data").rglob("*") if (root / "data").is_dir() else []:
        if not path.is_file():
            continue
        relative = path.relative_to(root).as_posix()
        profile: dict = {
            "path": relative,
            "suffix": path.suffix.lower(),
            "size_bytes": path.stat().st_size,
        }
        if path.suffix.lower() in {".xlsx", ".xlsm"}:
            profile["workbook"] = _profile_workbook(path)
        elif path.suffix.lower() == ".csv":
            profile["table"] = _profile_csv(path)
        files.append(profile)
    return {"files": files}


def _profile_workbook(path: Path) -> dict:
    from openpyxl import load_workbook

    workbook = load_workbook(path, read_only=False, data_only=True)
    try:
        sheets: list[dict] = []
        for sheet in workbook.worksheets:
            effective_max_column = max(
                (
                    cell.column
                    for row in sheet.iter_rows()
                    for cell in row
                    if cell.value is not None
                ),
                default=0,
            )
            sampled_rows = [
                (row_number, list(values))
                for row_number, values in enumerate(
                    sheet.iter_rows(
                        min_row=1,
                        max_row=min(sheet.max_row or 0, 6),
                        max_col=effective_max_column,
                        values_only=True,
                    ),
                    start=1,
                )
                if any(value is not None for value in values)
            ]
            header_row = sampled_rows[0][0] if sampled_rows else 0
            header_values = sampled_rows[0][1] if sampled_rows else []
            columns = [str(value) for value in header_values]
            merged_ranges = [
                merged
                for merged in sheet.merged_cells.ranges
                if merged.min_col <= effective_max_column
            ]
            vertical_group_ranges = [
                merged
                for merged in merged_ranges
                if merged.min_col == merged.max_col
                and merged.max_row > merged.min_row
                and merged.min_row > header_row
                and sheet.cell(merged.min_row, merged.min_col).value is not None
            ]
            forward_fill_columns = list(
                dict.fromkeys(
                    columns[merged.min_col - 1]
                    for merged in vertical_group_ranges
                    if merged.min_col <= len(columns)
                )
            )
            sample_rows: list[list] = []
            for row_number, raw_values in sampled_rows[1:5]:
                values = list(raw_values)
                for merged in vertical_group_ranges:
                    column_index = merged.min_col - 1
                    if merged.min_row <= row_number <= merged.max_row and values[column_index] is None:
                        values[column_index] = sheet.cell(merged.min_row, merged.min_col).value
                sample_rows.append([_json_safe_cell(value) for value in values])
            sheets.append(
                {
                    "name": sheet.title,
                    "max_row": sheet.max_row,
                    "max_column": sheet.max_column,
                    "effective_max_column": effective_max_column,
                    "ignored_trailing_empty_columns": max(
                        (sheet.max_column or 0) - effective_max_column,
                        0,
                    ),
                    "columns": columns,
                    "merged_ranges": [str(merged) for merged in merged_ranges],
                    "forward_fill_columns": forward_fill_columns,
                    "sample_rows": sample_rows,
                }
            )
        return {"sheets": sheets}
    finally:
        workbook.close()


def _profile_csv(path: Path) -> dict:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.reader(handle)
        rows = []
        for index, row in enumerate(reader):
            if index >= 6:
                break
            rows.append(row)
    return {
        "columns": rows[0] if rows else [],
        "sample_rows": rows[1:5],
    }


def _json_safe_cell(value):
    if value is None or isinstance(value, (str, int, float, bool)):
        return value
    return str(value)


def _is_positive_review_value(value) -> bool:
    text = str(value or "").strip().lower()
    positives = {
        "passed",
        "pass",
        "yes",
        "true",
        "consistent",
        "dimensionless",
        "not_applicable",
        "not applicable",
        "n/a",
    }
    return text in positives or any(
        text.startswith(f"{positive};") or text.startswith(f"{positive}:")
        for positive in positives
    )


def _write_model(path: Path, model) -> None:
    path.write_text(
        json.dumps(model.model_dump(mode="json"), ensure_ascii=False, indent=2),
        encoding="utf-8",
    )


def _read_json(path: Path) -> dict:
    payload = json.loads(path.read_text(encoding="utf-8"))
    return payload if isinstance(payload, dict) else {}


def _read_json_if_exists(path: Path) -> dict:
    return _read_json(path) if path.is_file() else {}


def _enforce_verified_guidance_evidence(
    report: VerificationReport,
    *,
    approved_model: dict,
    parameter_registry: dict,
    result_registry: dict,
) -> VerificationReport:
    if report.status != "verified":
        return report

    issues: list[str] = []
    parameters = parameter_registry.get("parameters")
    if not isinstance(parameters, list) or not parameters:
        issues.append("verified guidance evidence missing: parameter evidence")
    verified_results = result_registry.get("verified_results")
    if not isinstance(verified_results, list) or not verified_results:
        issues.append("verified guidance evidence missing: numerical results")

    check_groups = {
        "input checks": report.input_checks,
        "equation checks": report.equation_checks,
        "unit checks": report.unit_checks,
        "constraint checks": report.constraint_checks,
        "residual checks": report.residual_checks,
        "sensitivity checks": report.sensitivity_checks,
    }
    for label, checks in check_groups.items():
        if not checks:
            issues.append(f"verified guidance evidence missing: {label}")
        elif any(str(check.get("status")) != "passed" for check in checks):
            issues.append(f"verified guidance evidence failed: {label}")

    model = approved_model.get("approved_model")
    model = model if isinstance(model, dict) else {}
    subproblems = model.get("subproblem_models")
    if not isinstance(subproblems, list) or not subproblems:
        issues.append("verified guidance evidence missing: subproblem modeling details")
    else:
        for subproblem in subproblems:
            if not _has_complete_subproblem_guidance_details(subproblem):
                issues.append("verified guidance evidence incomplete: subproblem modeling details")
                break

    if not issues:
        return report
    return report.model_copy(
        update={
            "status": "blocked",
            "blocking_issues": [*report.blocking_issues, *issues],
        }
    )


def _has_complete_subproblem_guidance_details(subproblem: object) -> bool:
    if not isinstance(subproblem, dict):
        return False
    execution_status = subproblem.get("execution_status")
    if execution_status == "solvable":
        return all(
            isinstance(subproblem.get(field), list) and bool(subproblem[field])
            for field in ("equations", "parameters", "solve_steps")
        )
    if execution_status == "blocked":
        return (
            bool(str(subproblem.get("blocking_reason") or "").strip())
            and isinstance(subproblem.get("missing_inputs"), list)
            and bool(subproblem["missing_inputs"])
            and bool(str(subproblem.get("recovery_action") or "").strip())
        )
    return False


def _operator_evidence_issues(
    *,
    root: Path,
    executed_node_ids: set[str],
    parameter_registry: dict,
    result_registry: dict,
    figure_registry: dict,
) -> tuple[list[str], list[dict]]:
    issues: list[str] = []
    parameters = parameter_registry.get("parameters")
    if not isinstance(parameters, list):
        issues.append("parameter_registry.parameters must be a list")
        parameters = []
    for parameter in parameters:
        if not isinstance(parameter, dict):
            issues.append("parameter registry contains a non-object entry")
            continue
        source_node = str(parameter.get("source_node") or "")
        if source_node not in executed_node_ids:
            issues.append(
                f"parameter {parameter.get('parameter_id')} has an unexecuted source node"
            )

    verified_results = result_registry.get("verified_results")
    blocked_results = result_registry.get("blocked_results")
    if not isinstance(verified_results, list):
        issues.append("result_registry.verified_results must be a list")
        verified_results = []
    if not isinstance(blocked_results, list):
        issues.append("result_registry.blocked_results must be a list")
    result_ids: set[str] = set()
    for result in verified_results:
        if not isinstance(result, dict):
            issues.append("result registry contains a non-object verified entry")
            continue
        result_id = str(result.get("id") or "")
        result_ids.add(result_id)
        source_node = str(result.get("source_node") or "")
        if source_node not in executed_node_ids:
            issues.append(f"result {result_id} has an unexecuted source node")
        evidence_ids = result.get("evidence_ids")
        if not isinstance(evidence_ids, list) or not evidence_ids:
            issues.append(f"result {result_id} has no evidence node IDs")
        else:
            unknown = sorted(set(map(str, evidence_ids)) - executed_node_ids)
            if unknown:
                issues.append(
                    f"result {result_id} references unexecuted evidence nodes: {', '.join(unknown)}"
                )

    figures = figure_registry.get("figures")
    if not isinstance(figures, list):
        issues.append("figure_registry.figures must be a list")
        figures = []
    for figure in figures:
        if not isinstance(figure, dict):
            issues.append("figure registry contains a non-object entry")
            continue
        path = str(figure.get("path") or "")
        if not path or not (root / path).is_file():
            issues.append(f"registered figure was not generated: {path or '<missing path>'}")
        linked_result_ids = figure.get("linked_result_ids")
        if not isinstance(linked_result_ids, list) or not linked_result_ids:
            issues.append(f"figure {path or '<missing path>'} has no linked result IDs")
            continue
        unknown_ids = set(map(str, linked_result_ids)) - result_ids
        if unknown_ids:
            issues.append(
                f"figure has invalid result IDs for {path}: {', '.join(sorted(unknown_ids))}"
            )

    return issues, [figure for figure in figures if isinstance(figure, dict)]


def _build_blocked_guidance_artifacts(
    root: Path,
    approved_path: Path,
) -> tuple[VerificationReport, dict, dict, dict]:
    approved_model = ApprovedModelIR.model_validate_json(
        approved_path.read_text(encoding="utf-8")
    )
    blocked_subproblems = [
        subproblem
        for subproblem in approved_model.approved_model_ir.subproblems
        if subproblem.execution_status == "blocked"
    ]
    issues = [subproblem.blocking_reason for subproblem in blocked_subproblems]

    parameters = {
        "parameters": [],
        "summary": {"total_count": 0, "blocked_count": 0},
    }
    results = {
        "verified_results": [],
        "blocked_results": [
            {
                "id": f"blocked_{subproblem.subproblem_id.replace('.', '_')}",
                "message": subproblem.blocking_reason,
                "reason": "approved_model_ir_blocked",
                "subproblem_id": subproblem.subproblem_id,
                "missing_inputs": subproblem.missing_inputs,
                "recovery_action": subproblem.recovery_action,
                "expected_outputs": subproblem.expected_outputs,
                "source_data": [approved_model.model_ir_ref],
            }
            for subproblem in blocked_subproblems
        ],
        "summary": {
            "coverage_status": "blocked",
            "verified_count": 0,
            "blocked_count": len(issues),
        },
    }
    report = VerificationReport(
        model_id=approved_model.model_id,
        status="blocked",
        blocking_issues=issues,
        artifact_refs={
            "approved_model": approved_path.name,
            "model_review": approved_model.review_ref,
            "parameter_registry": "parameter_registry.json",
            "result_registry": "result_registry.json",
            "figures": [],
        },
    )
    _write_model(root / "verification_report.json", report)
    (root / "parameter_registry.json").write_text(
        json.dumps(parameters, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    (root / "result_registry.json").write_text(
        json.dumps(results, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    return report, approved_model.model_dump(mode="json"), parameters, results

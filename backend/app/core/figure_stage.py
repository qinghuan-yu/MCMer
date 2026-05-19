"""Figure stage gate and deterministic rendering helpers extracted from workflow."""

import json
import csv
import re
from pathlib import Path
from typing import Any

from app.artifacts.renderers import (
    render_algorithm_flowchart,
    render_comparison_bar,
    render_distance_time_curve,
    render_fitting_curve,
    render_geometry_schematic,
    render_peak_valley_annotation,
    render_residual_plot,
    render_spectrum_curve,
    render_timeline_schematic,
    render_trajectory_shielding_2d,
)
from app.artifacts.figure_plan import FigurePlanBuilder
from app.artifacts.registry import ArtifactRegistry


class FigureStage:
    _FIGURE_KEYWORDS = re.compile(
        r"(?:图表|作图|绘图|制图|画图|插图|示意图|流程图|架构图|拓扑图|"
        r"折线图|散点图|柱状图|饼图|箱线图|热力图|等高线图|雷达图|直方图|"
        r"figure|chart|plot|graph|diagram|visualization|visuali[sz]e|"
        r"scatter|histogram|heatmap|contour|boxplot|bar\s*chart|pie\s*chart|"
        r"折线|柱状|散点|热力|等高线|箱线|雷达|可视化)",
        re.IGNORECASE,
    )

    @staticmethod
    def ensure_figure_requests_in_solve_spec(
        solve_spec: dict[str, object],
        chart_language: str,
    ) -> list[dict[str, object]]:
        existing = solve_spec.get("figure_requests") if isinstance(solve_spec, dict) else None
        if isinstance(existing, list) and any(isinstance(item, dict) for item in existing):
            normalized = [item for item in existing if isinstance(item, dict)]
            solve_spec["figure_requests"] = normalized
            return normalized

        requests: list[dict[str, object]] = []
        seen_ids: set[str] = set()
        subproblems_raw = solve_spec.get("subproblems", []) if isinstance(solve_spec, dict) else []
        subproblems = [item for item in subproblems_raw if isinstance(item, dict)] if isinstance(subproblems_raw, list) else []

        def _default_required_data_for_role(role: str) -> list[str]:
            if role in {"geometry_schematic", "timeline_schematic", "trajectory_overview", "optimization_variable_diagram"}:
                return ["result_registry.json"]
            if role == "trajectory_shielding":
                return [
                    "missile_trajectory.json",
                    "cloud_center_trajectory.json",
                    "shielding_intervals.json",
                ]
            if role == "distance_time_curve":
                return ["distance_time_series.csv"]
            if role == "optimization_convergence":
                return ["optimization_history.csv"]
            if role == "sensitivity_analysis":
                return ["sensitivity_results.csv"]
            return ["key_result_series.csv"]

        def _infer_figure_role(text: str) -> str:
            lowered = text.lower()
            if any(token in lowered for token in ["轨迹", "trajectory"]):
                return "trajectory_shielding"
            if any(token in lowered for token in ["收敛", "convergence", "优化"]):
                return "optimization_convergence"
            if any(token in lowered for token in ["距离", "distance", "时间", "time"]):
                return "distance_time_curve"
            if any(token in lowered for token in ["敏感性", "sensitivity"]):
                return "sensitivity_analysis"
            return "result_visualization"

        def _append_request(
            subproblem_id: str,
            role: str,
            required: bool,
            expected_text: str,
            figure_kind: str,
        ) -> None:
            request_id = f"fig_{subproblem_id}_{role}"
            if request_id in seen_ids:
                return
            seen_ids.add(request_id)
            requests.append(
                {
                    "id": request_id,
                    "subproblem_id": subproblem_id,
                    "required": required,
                    "role": role,
                    "semantic_role": role,
                    "figure_kind": figure_kind,
                    "expected_content": [expected_text],
                    "required_data": _default_required_data_for_role(role),
                    "language": chart_language,
                    "chart_language": chart_language,
                    "must_include_in_paper": bool(required and figure_kind == "explanatory_figure"),
                }
            )

        geometry_keywords = re.compile(r"导弹|无人机|目标|烟幕|遮蔽|视线|几何|geometry|line\s*of\s*sight", re.IGNORECASE)
        trajectory_keywords = re.compile(r"轨迹|trajectory|path", re.IGNORECASE)
        timeline_keywords = re.compile(r"投放|起爆|有效期|时间区间|时刻|timeline|time\s*window", re.IGNORECASE)
        optimization_keywords = re.compile(r"优化变量|约束|目标函数|优化|optimization|objective|constraint", re.IGNORECASE)

        for index, subproblem in enumerate(subproblems, start=1):
            text = json.dumps(subproblem, ensure_ascii=False)
            if not FigureStage._FIGURE_KEYWORDS.search(text):
                continue
            subproblem_id = str(subproblem.get("id") or f"problem_{index}").strip() or f"problem_{index}"
            objective = str(subproblem.get("objective") or "").strip() or subproblem_id

            if geometry_keywords.search(text):
                _append_request(subproblem_id, "geometry_schematic", True, f"{objective} 几何关系示意图", "explanatory_figure")
            if trajectory_keywords.search(text):
                _append_request(subproblem_id, "trajectory_overview", True, f"{objective} 轨迹总览示意图", "explanatory_figure")
            if timeline_keywords.search(text):
                _append_request(subproblem_id, "timeline_schematic", True, f"{objective} 时间轴示意图", "explanatory_figure")
            if optimization_keywords.search(text):
                _append_request(subproblem_id, "optimization_variable_diagram", False, f"{objective} 优化变量关系图", "explanatory_figure")

            role = _infer_figure_role(text)
            _append_request(subproblem_id, role, True, objective, "result_figure")

        solve_spec["figure_requests"] = requests
        return requests

    @staticmethod
    def normalize_workflow_issues(figure_plan: dict[str, object] | None) -> list[dict[str, object]]:
        if not isinstance(figure_plan, dict):
            return []
        raw = figure_plan.get("workflow_issues", [])
        if not isinstance(raw, list):
            return []
        issues: list[dict[str, object]] = []
        for item in raw:
            if not isinstance(item, dict):
                continue
            issues.append(
                {
                    "code": str(item.get("code") or "").strip(),
                    "severity": str(item.get("severity") or "").strip() or "warning",
                    "stage": str(item.get("stage") or "").strip() or "figure_plan",
                    "reason": str(item.get("reason") or "").strip() or "unspecified",
                    "request_id": str(item.get("request_id") or "").strip(),
                    "action": str(item.get("action") or "").strip(),
                }
            )
        return issues

    @staticmethod
    def workflow_issue_summary(issues: list[dict[str, object]]) -> dict[str, Any]:
        summary: dict[str, Any] = {
            "count": 0,
            "by_code": {},
            "by_stage": {},
            "top_reasons": {},
        }
        by_code: dict[str, int] = {}
        by_stage: dict[str, int] = {}
        top_reasons: dict[str, int] = {}
        valid_count = 0
        for item in issues:
            if not isinstance(item, dict):
                continue
            valid_count += 1
            code = str(item.get("code") or "UNKNOWN").strip() or "UNKNOWN"
            stage = str(item.get("stage") or "unknown").strip() or "unknown"
            reason = str(item.get("reason") or "unspecified").strip() or "unspecified"
            by_code[code] = by_code.get(code, 0) + 1
            by_stage[stage] = by_stage.get(stage, 0) + 1
            top_reasons[reason] = top_reasons.get(reason, 0) + 1

        summary["count"] = valid_count
        summary["by_code"] = by_code
        summary["by_stage"] = by_stage
        summary["top_reasons"] = dict(
            sorted(top_reasons.items(), key=lambda kv: kv[1], reverse=True)[:5]
        )
        return summary

    @staticmethod
    def reevaluate_solve_spec_requests(
        solve_spec: dict[str, object],
        chart_language: str,
        work_dir: str,
    ) -> list[dict[str, object]]:
        if not isinstance(solve_spec, dict):
            return []
        current_requests = solve_spec.get("figure_requests", [])
        if not isinstance(current_requests, list) or not current_requests:
            return []

        reevaluated = FigurePlanBuilder.reevaluate_requests(
            requests=[item for item in current_requests if isinstance(item, dict)],
            chart_language=chart_language,
            work_dir=work_dir,
        )
        if not reevaluated:
            return []

        solve_spec["figure_requests"] = reevaluated
        solve_spec["required_feasible_count"] = sum(
            1 for item in reevaluated if isinstance(item, dict) and bool(item.get("required", True))
        )
        solve_spec["blocked_count"] = sum(
            1 for item in reevaluated if isinstance(item, dict) and str(item.get("status") or "") == "blocked"
        )
        return [item for item in reevaluated if isinstance(item, dict)]

    @staticmethod
    def required_figure_requests(solve_spec: dict[str, object]) -> list[dict[str, object]]:
        requests = solve_spec.get("figure_requests", []) if isinstance(solve_spec, dict) else []
        if not isinstance(requests, list):
            return []
        return [
            req
            for req in requests
            if isinstance(req, dict)
            and bool(req.get("required", True))
            and str(req.get("status") or "").strip().lower() != "blocked"
            and req.get("feasible", True) is not False
        ]

    @staticmethod
    def figure_requests_for_subproblem(
        solve_spec: dict[str, object],
        subproblem_id: str,
    ) -> list[dict[str, object]]:
        requests = solve_spec.get("figure_requests", []) if isinstance(solve_spec, dict) else []
        if not isinstance(requests, list):
            return []
        sid = str(subproblem_id or "").strip()
        if not sid:
            return []
        scoped: list[dict[str, object]] = []
        for item in requests:
            if not isinstance(item, dict):
                continue
            if str(item.get("subproblem_id") or item.get("problem_section") or "").strip() == sid:
                scoped.append(item)
        return scoped

    @staticmethod
    def figure_data_protocol_contract(requests: list[dict[str, object]]) -> str:
        if not requests:
            return ""

        lines: list[str] = []
        lines.append("## Figure data protocol (deterministic renderers)")
        lines.append("当 figure_request 包含以下语义角色时，Solver 必须先产出结构化数据文件，再产图：")

        def _req_data_files(req: dict[str, object]) -> list[str]:
            required_data = req.get("required_data", [])
            from_required = [str(item).strip() for item in required_data if str(item).strip()] if isinstance(required_data, list) else []
            depends_on = req.get("depends_on", {}) if isinstance(req.get("depends_on"), dict) else {}
            dep_files = depends_on.get("data_files", []) if isinstance(depends_on.get("data_files"), list) else []
            from_depends = [str(item).strip() for item in dep_files if str(item).strip()]
            merged = [*from_required, *from_depends]
            merged = [item for item in merged if item and item != "result_registry.json"]
            return list(dict.fromkeys(merged))

        for req in requests:
            if not isinstance(req, dict):
                continue
            role = str(req.get("semantic_role") or req.get("role") or "").strip().lower()
            if role not in {"fitting_curve", "residual_plot", "comparison_bar"}:
                continue
            req_id = str(req.get("id") or "").strip() or "unknown_request"
            linked_ids = req.get("linked_result_ids", [])
            linked = [str(item).strip() for item in linked_ids if str(item).strip()] if isinstance(linked_ids, list) else []
            data_files = _req_data_files(req)
            data_file_hint = data_files[0] if data_files else "data/<subproblem>_data.csv"

            if role == "fitting_curve":
                lines.append(
                    f"- {req_id} fitting_curve: 生成 CSV `{data_file_hint}`，至少包含列 x, observed, fitted；"
                    f"并把对应结果写入 key_results（建议 id: {linked[0] if linked else 'q?_fitting_result'}）。"
                )
            elif role == "residual_plot":
                lines.append(
                    f"- {req_id} residual_plot: 生成 CSV `{data_file_hint}`，至少包含列 x, residual；"
                    f"并把对应结果写入 key_results（建议 id: {linked[0] if linked else 'q?_residual_result'}）。"
                )
            else:
                lines.append(
                    f"- {req_id} comparison_bar: 生成 CSV `{data_file_hint}`，优先列 method, metric, value（可兼容 category, value）；"
                    f"并把对应结果写入 key_results（建议 id: {linked[0] if linked else 'q?_comparison_result'}）。"
                )

        if len(lines) <= 2:
            return ""
        lines.append("未按列协议产出数据文件时，对应图将被降级为 blocked，不进入 required。")
        return "\n".join(lines) + "\n"

    @staticmethod
    def apply_figure_plan_to_solve_spec(
        solve_spec: dict[str, object],
        figure_plan: dict[str, object],
    ) -> None:
        if not isinstance(solve_spec, dict) or not isinstance(figure_plan, dict):
            return
        requires_figures_by_problem = bool(
            figure_plan.get("requires_figures_by_problem", figure_plan.get("requires_figures"))
        )
        requires_result_by_problem = bool(
            figure_plan.get("requires_result_figures_by_problem", figure_plan.get("requires_result_figures"))
        )
        requires_explanatory_by_problem = bool(
            figure_plan.get("requires_explanatory_figures_by_problem", figure_plan.get("requires_explanatory_figures"))
        )

        solve_spec["requires_figures"] = requires_figures_by_problem
        solve_spec["requires_result_figures"] = requires_result_by_problem
        solve_spec["requires_explanatory_figures"] = requires_explanatory_by_problem
        solve_spec["requires_figures_by_problem"] = requires_figures_by_problem
        solve_spec["requires_result_figures_by_problem"] = requires_result_by_problem
        solve_spec["requires_explanatory_figures_by_problem"] = requires_explanatory_by_problem
        solve_spec["required_feasible_count"] = int(figure_plan.get("required_feasible_count", figure_plan.get("required_count", 0)) or 0)
        solve_spec["blocked_count"] = int(figure_plan.get("blocked_count", 0) or 0)
        solve_spec["expected_figures"] = requires_figures_by_problem

    @staticmethod
    def subproblem_figure_mode(
        subproblem: dict[str, object],
        task_expected_figures: bool,
    ) -> str:
        if not task_expected_figures:
            return "none"
        text = json.dumps(subproblem, ensure_ascii=False).lower()
        if FigureStage._FIGURE_KEYWORDS.search(text):
            return "paper_required"
        return "none"

    @staticmethod
    def build_bundle_snapshot(art_registry: ArtifactRegistry) -> dict[str, object]:
        figure_bundle = art_registry.build_figure_bundle()
        writer_images = [
            str(item.get("path") or "").strip()
            for item in figure_bundle.available_figures
            if str(item.get("path") or "").strip()
        ]
        return {
            "figure_bundle": figure_bundle,
            "writer_images": writer_images,
            "figure_bundle_dict": figure_bundle.to_dict(),
        }

    @staticmethod
    def repair_eligible_missing_reason(reason: str) -> bool:
        lowered = str(reason or "").strip().lower()
        if not lowered:
            return True
        blocked_prefixes = (
            "unsupported_deterministic_role",
            "required_data_missing",
            "insufficient_data_series",
            "insufficient_semantic_required_data_contract",
            "missing_fitting_columns",
            "missing_residual_columns",
            "missing_comparison_columns",
            "missing_source_data",
            "missing_verified_result",
            "linked_result_not_verified",
            "placeholder_linked_result_ids",
            "pending_result_binding",
            "unsupported_renderer",
        )
        return not any(lowered.startswith(prefix) for prefix in blocked_prefixes)

    @staticmethod
    def repair_eligible_request(target_req: dict[str, object] | None, missing_reason: str) -> bool:
        if not FigureStage.repair_eligible_missing_reason(missing_reason):
            return False
        if not isinstance(target_req, dict):
            return True
        if str(target_req.get("status") or "").strip().lower() == "blocked":
            return False
        if target_req.get("feasible", True) is False:
            return False
        if str(target_req.get("blocked_reason") or "").strip():
            return False
        if str(target_req.get("result_binding_status") or "").strip().lower() == "pending_result_binding":
            return False
        return True

    @staticmethod
    def missing_reason_summary(missing_requests: list[dict[str, object]]) -> dict[str, int]:
        summary = {
            "unsupported_renderer": 0,
            "missing_verified_result": 0,
            "missing_source_data": 0,
            "repair_budget_exceeded": 0,
            "unbound_semantic_artifact": 0,
            "other": 0,
        }
        for item in missing_requests:
            reason = str(item.get("reason") or "").strip().lower()
            if reason.startswith("unsupported_deterministic_role") or reason.startswith("unsupported_renderer"):
                summary["unsupported_renderer"] += 1
            elif (
                "linked_result_not_verified" in reason
                or "placeholder_linked_result_ids" in reason
                or "missing_verified_result" in reason
            ):
                summary["missing_verified_result"] += 1
            elif (
                reason.startswith("required_data_missing")
                or "missing_source_data" in reason
                or "insufficient_data_series" in reason
                or "missing_fitting_columns" in reason
                or "missing_residual_columns" in reason
                or "missing_comparison_columns" in reason
            ):
                summary["missing_source_data"] += 1
            elif "budget" in reason or "tool calls exceeded" in reason:
                summary["repair_budget_exceeded"] += 1
            elif "no semantic_verified artifact" in reason:
                summary["unbound_semantic_artifact"] += 1
            else:
                summary["other"] += 1
        return summary

    @staticmethod
    def quality_gate_before_writing(
        result_registry: dict[str, object],
        expected_figures: bool,
        required_requests: list[dict[str, object]],
        satisfied_requests: list[dict[str, object]],
        missing_requests: list[dict[str, object]],
        workflow_mode: str,
        requires_result_figures: bool = False,
        requires_explanatory_figures: bool = False,
        result_figure_count: int = 0,
        explanatory_figure_count: int = 0,
        requires_figures_by_problem: bool = False,
        blocked_requests: list[dict[str, object]] | None = None,
    ) -> tuple[bool, str]:
        """Return (passed, reason). If passed is False, reason explains why."""
        summary = result_registry.get("summary", {}) if isinstance(result_registry, dict) else {}
        verified_count = summary.get("verified_count", 0) if isinstance(summary, dict) else 0
        blocked_count = summary.get("blocked_count", 0) if isinstance(summary, dict) else 0

        if verified_count == 0:
            return False, (
                f"verified_count=0（blocked_count={blocked_count}）。"
                "无法输出正式论文，因为没有任何经过验证的数值结果。"
                "请检查求解阶段是否正确产出 key_results 并标记为 verified。"
            )

        if expected_figures and required_requests and missing_requests:
            missing_ids = [str(item.get("figure_request_id") or "").strip() for item in missing_requests]
            missing_ids = [item for item in missing_ids if item]
            reason_summary = FigureStage.missing_reason_summary(missing_requests)
            return False, (
                "任务预期有图表（expected_figures=true），但未满足全部 required figure_requests。"
                f"missing={missing_ids}。"
                f"missing_by_reason={reason_summary}。"
                "请检查图表阶段是否为每个 required request 产出 semantic_verified=true 且绑定 figure_request_id 的图。"
            )

        if expected_figures and not required_requests:
            blocked = blocked_requests if isinstance(blocked_requests, list) else []
            if requires_figures_by_problem and blocked:
                blocked_by_reason: dict[str, int] = {}
                for item in blocked:
                    if not isinstance(item, dict):
                        continue
                    reason = str(item.get("blocked_reason") or "unspecified").strip() or "unspecified"
                    blocked_by_reason[reason] = blocked_by_reason.get(reason, 0) + 1
                return False, (
                    "任务按题意需要图表（requires_figures_by_problem=true），但当前没有可行 required figure。"
                    f"blocked_required_by_problem_count={len(blocked)}。"
                    f"blocked_by_reason={blocked_by_reason}。"
                    "请补齐 renderer 或数据绑定后再将相关请求升级为 required。"
                )
            return False, (
                "任务预期有图表（expected_figures=true），但 solve_spec 未提供 figure_requests。"
                "请在求解规格阶段生成 required figure_requests。"
            )

        if expected_figures and required_requests and not satisfied_requests:
            return False, (
                "任务预期有图表（expected_figures=true），但没有任何 required figure_request 被满足。"
                "请确保产物包含 figure_request_id 且 semantic_verified=true。"
            )

        if requires_explanatory_figures and explanatory_figure_count <= 0:
            return False, (
                "任务要求 explanatory figures，但未生成任何说明图。"
                "请生成 geometry/timeline/trajectory/optimization 语义图并满足 figure_request。"
            )

        if requires_result_figures and result_figure_count <= 0:
            return False, (
                "任务要求 result figures，但未生成任何结果图。"
                "请生成绑定 verified_results 的结果图。"
            )

        return True, ""

    @staticmethod
    def has_semantic_required_data_contract(role: str, required_data: list[str]) -> bool:
        normalized = [str(item).strip().lower() for item in required_data if str(item).strip()]
        if not normalized:
            return False
        if len(normalized) == 1 and normalized[0] == "result_registry.json":
            return False

        merged = " ".join(normalized)
        if role == "trajectory_shielding":
            return "trajectory" in merged and ("shield" in merged or "cloud" in merged)
        if role == "distance_time_curve":
            return "distance" in merged and "time" in merged
        return len(normalized) >= 1

    @staticmethod
    def coerce_numeric_value(value: object) -> float | None:
        if isinstance(value, (int, float)) and not isinstance(value, bool):
            return float(value)
        text = str(value or "")
        match = re.search(r"[-+]?(?:\d+(?:\.\d*)?|\.\d+)(?:[eE][-+]?\d+)?", text)
        if not match:
            return None
        try:
            return float(match.group(0))
        except ValueError:
            return None

    @staticmethod
    def deterministic_series_from_registry(
        result_registry: dict[str, object],
        required_data: list[str],
    ) -> tuple[list[float], list[float], list[str]]:
        verified = result_registry.get("verified_results", []) if isinstance(result_registry, dict) else []
        if not isinstance(verified, list):
            return [], [], []

        values: list[float] = []
        matched_sources: set[str] = set()
        required_set = {str(item).strip() for item in required_data if str(item).strip()}

        for entry in verified:
            if not isinstance(entry, dict):
                continue
            entry_sources = [str(item).strip() for item in (entry.get("source_data", []) or []) if str(item).strip()]
            if required_set and not (required_set.intersection(set(entry_sources))):
                continue
            value = FigureStage.coerce_numeric_value(entry.get("value"))
            if value is None:
                continue
            result_id = str(entry.get("id") or "").strip()
            if not result_id:
                continue
            values.append(value)
            for src in entry_sources:
                matched_sources.add(src)
            if len(values) >= 24:
                break

        x_values = [float(index + 1) for index in range(len(values))]
        return x_values, values, sorted(matched_sources)

    @staticmethod
    def deterministic_series_from_data_files(
        work_dir: str,
        data_files: list[str],
    ) -> tuple[list[float], list[float], list[str]]:
        root = Path(work_dir)
        candidates: list[tuple[Path, str]] = []
        for raw in data_files:
            rel = str(raw or "").strip()
            if not rel:
                continue
            path = root / rel
            if path.exists() and path.is_file():
                candidates.append((path, rel.replace("\\", "/")))

        if not candidates:
            return [], [], []

        def _to_float(v: object) -> float | None:
            try:
                text = str(v).strip()
                if not text:
                    return None
                return float(text)
            except Exception:
                return None

        def _pick_columns(column_map: dict[str, list[float]]) -> tuple[list[float], list[float]]:
            if not column_map:
                return [], []
            names = list(column_map.keys())

            x_candidates = [
                key for key in names if any(token in key.lower() for token in ["wavenumber", "波数", "time", "时间", "x"])
            ]
            y_candidates = [
                key for key in names if any(token in key.lower() for token in ["reflectance", "反射率", "intensity", "value", "y"])
            ]

            x_name = x_candidates[0] if x_candidates else ""
            y_name = y_candidates[0] if y_candidates else ""

            if not y_name:
                for name in names:
                    if name != x_name and len(column_map.get(name, [])) >= 2:
                        y_name = name
                        break
            if not x_name:
                x_name = next((name for name in names if name != y_name and len(column_map.get(name, [])) >= 2), "")

            if y_name:
                y_values = column_map.get(y_name, [])
                if x_name:
                    x_values = column_map.get(x_name, [])
                    if len(x_values) == len(y_values) and len(y_values) >= 2:
                        return x_values, y_values
                if len(y_values) >= 2:
                    return [float(i + 1) for i in range(len(y_values))], y_values
            return [], []

        for path, rel in candidates:
            suffix = path.suffix.lower()
            if suffix == ".csv":
                try:
                    with path.open("r", encoding="utf-8", newline="") as fp:
                        reader = csv.DictReader(fp)
                        if not reader.fieldnames:
                            continue
                        column_map: dict[str, list[float]] = {str(name): [] for name in reader.fieldnames if str(name).strip()}
                        for row in reader:
                            for name in list(column_map.keys()):
                                val = _to_float((row or {}).get(name, ""))
                                if val is not None:
                                    column_map[name].append(val)
                    x_values, y_values = _pick_columns(column_map)
                    if len(x_values) >= 2 and len(y_values) >= 2:
                        return x_values, y_values, [rel]
                except Exception:
                    continue

            if suffix in {".xlsx", ".xls"}:
                try:
                    from openpyxl import load_workbook  # type: ignore

                    wb = load_workbook(path, data_only=True, read_only=True)
                    ws = wb.active
                    rows = ws.iter_rows(min_row=1, max_row=2000, values_only=True)
                    header = next(rows, None)
                    if not header:
                        continue
                    names = [str(item or "").strip() for item in header]
                    column_map: dict[str, list[float]] = {name: [] for name in names if name}
                    for row in rows:
                        for idx, name in enumerate(names):
                            if not name:
                                continue
                            val = _to_float(row[idx] if row and idx < len(row) else "")
                            if val is not None:
                                column_map[name].append(val)
                    x_values, y_values = _pick_columns(column_map)
                    if len(x_values) >= 2 and len(y_values) >= 2:
                        return x_values, y_values, [rel]
                except Exception:
                    continue

        return [], [], []

    @staticmethod
    def deterministic_named_columns_from_data_files(
        work_dir: str,
        data_files: list[str],
    ) -> tuple[dict[str, list[float]], list[str]]:
        root = Path(work_dir)
        candidates: list[tuple[Path, str]] = []
        for raw in data_files:
            rel = str(raw or "").strip()
            if not rel:
                continue
            path = root / rel
            if path.exists() and path.is_file():
                candidates.append((path, rel.replace("\\", "/")))

        def _to_float(v: object) -> float | None:
            try:
                text = str(v).strip()
                if not text:
                    return None
                return float(text)
            except Exception:
                return None

        for path, rel in candidates:
            suffix = path.suffix.lower()
            if suffix == ".csv":
                try:
                    with path.open("r", encoding="utf-8", newline="") as fp:
                        reader = csv.DictReader(fp)
                        if not reader.fieldnames:
                            continue
                        column_map: dict[str, list[float]] = {str(name): [] for name in reader.fieldnames if str(name).strip()}
                        for row in reader:
                            for name in list(column_map.keys()):
                                val = _to_float((row or {}).get(name, ""))
                                if val is not None:
                                    column_map[name].append(val)
                    if any(len(values) >= 2 for values in column_map.values()):
                        return column_map, [rel]
                except Exception:
                    continue
            if suffix in {".xlsx", ".xls"}:
                try:
                    from openpyxl import load_workbook  # type: ignore

                    wb = load_workbook(path, data_only=True, read_only=True)
                    ws = wb.active
                    rows = ws.iter_rows(min_row=1, max_row=3000, values_only=True)
                    header = next(rows, None)
                    if not header:
                        continue
                    names = [str(item or "").strip() for item in header]
                    column_map: dict[str, list[float]] = {name: [] for name in names if name}
                    for row in rows:
                        for idx, name in enumerate(names):
                            if not name:
                                continue
                            val = _to_float(row[idx] if row and idx < len(row) else "")
                            if val is not None:
                                column_map[name].append(val)
                    if any(len(values) >= 2 for values in column_map.values()):
                        return column_map, [rel]
                except Exception:
                    continue

        return {}, []

    @staticmethod
    def extract_named_series(columns: dict[str, list[float]], aliases: list[str]) -> list[float]:
        lowered_aliases = [alias.lower() for alias in aliases]
        for name, values in columns.items():
            lowered = name.lower()
            if any(token in lowered for token in lowered_aliases) and len(values) >= 2:
                return values
        return []

    @staticmethod
    def render_required_requests_deterministically(
        work_dir: str,
        result_registry: dict[str, object],
        required_requests: list[dict[str, object]],
        chart_language: str,
    ) -> dict[str, object]:
        created_images: list[str] = []
        satisfied: list[dict[str, object]] = []
        missing: list[dict[str, object]] = []

        for req in required_requests:
            req_id = str(req.get("id") or "").strip()
            subproblem_id = str(req.get("subproblem_id") or "").strip()
            role = str(req.get("semantic_role") or req.get("role") or "").strip().lower()
            figure_kind = str(req.get("figure_kind") or "result_figure").strip().lower()
            expected_content = req.get("expected_content", [])
            required_data = req.get("required_data", [])
            required_data_list = [str(item).strip() for item in required_data if str(item).strip()] if isinstance(required_data, list) else []
            depends_on = req.get("depends_on", {}) if isinstance(req.get("depends_on"), dict) else {}
            model_objects = depends_on.get("model_objects", []) if isinstance(depends_on.get("model_objects"), list) else []
            formulas = depends_on.get("formulas", []) if isinstance(depends_on.get("formulas"), list) else []
            source_problem_refs = depends_on.get("source_problem_refs", []) if isinstance(depends_on.get("source_problem_refs"), list) else []

            if not req_id:
                continue
            if role not in {
                "trajectory_shielding",
                "distance_time_curve",
                "geometry_schematic",
                "physical_principle_schematic",
                "optical_path_diagram",
                "timeline_schematic",
                "trajectory_overview",
                "optimization_variable_diagram",
                "algorithm_flowchart",
                "spectrum_curve",
                "peak_valley_annotation",
                "fitting_curve",
                "residual_plot",
                "comparison_bar",
            }:
                missing.append({
                    "figure_request_id": req_id,
                    "required": True,
                    "satisfied": False,
                    "reason": f"unsupported_deterministic_role:{role or 'unknown'}",
                })
                continue

            if role in {"trajectory_shielding", "distance_time_curve"} and not FigureStage.has_semantic_required_data_contract(role, required_data_list):
                missing.append({
                    "figure_request_id": req_id,
                    "required": True,
                    "satisfied": False,
                    "reason": "insufficient_semantic_required_data_contract",
                })
                continue

            if required_data_list:
                missing_files = [src for src in required_data_list if not (Path(work_dir) / src).exists()]
                if missing_files and role not in {"spectrum_curve", "peak_valley_annotation"}:
                    missing.append({
                        "figure_request_id": req_id,
                        "required": True,
                        "satisfied": False,
                        "reason": f"required_data_missing:{missing_files}",
                    })
                    continue
                if role in {"spectrum_curve", "peak_valley_annotation"} and len(missing_files) == len(required_data_list):
                    missing.append({
                        "figure_request_id": req_id,
                        "required": True,
                        "satisfied": False,
                        "reason": f"required_data_missing:{missing_files}",
                    })
                    continue

            x_values: list[float] = []
            y_values: list[float] = []
            matched_sources: list[str] = []
            observed_values: list[float] = []
            fitted_values: list[float] = []
            categories: list[str] = []
            comparison_values: list[float] = []
            if role in {"trajectory_shielding", "distance_time_curve"}:
                x_values, y_values, matched_sources = FigureStage.deterministic_series_from_registry(
                    result_registry,
                    required_data_list,
                )
                if len(x_values) < 2:
                    missing.append({
                        "figure_request_id": req_id,
                        "required": True,
                        "satisfied": False,
                        "reason": "insufficient_verified_numeric_series",
                    })
                    continue

            if role in {"spectrum_curve", "peak_valley_annotation"}:
                x_values, y_values, matched_sources = FigureStage.deterministic_series_from_data_files(
                    work_dir,
                    required_data_list,
                )
                if len(x_values) < 2:
                    x_values, y_values, matched_sources = FigureStage.deterministic_series_from_registry(
                        result_registry,
                        required_data_list,
                    )
                if len(x_values) < 2:
                    missing.append({
                        "figure_request_id": req_id,
                        "required": True,
                        "satisfied": False,
                        "reason": "insufficient_data_series",
                    })
                    continue

            if role in {"fitting_curve", "residual_plot", "comparison_bar"}:
                columns, matched_sources = FigureStage.deterministic_named_columns_from_data_files(
                    work_dir,
                    required_data_list,
                )
                if not columns:
                    missing.append({
                        "figure_request_id": req_id,
                        "required": True,
                        "satisfied": False,
                        "reason": "missing_source_data",
                    })
                    continue
                if role == "fitting_curve":
                    x_values = FigureStage.extract_named_series(columns, ["x", "time", "波数", "wavenumber", "index"])
                    observed_values = FigureStage.extract_named_series(columns, ["observed", "obs", "measured", "真实", "观测"])
                    fitted_values = FigureStage.extract_named_series(columns, ["fitted", "fit", "pred", "预测", "拟合"])
                    if not x_values and observed_values:
                        x_values = [float(i + 1) for i in range(len(observed_values))]
                    if not x_values or not observed_values or not fitted_values:
                        missing.append({
                            "figure_request_id": req_id,
                            "required": True,
                            "satisfied": False,
                            "reason": "missing_fitting_columns",
                        })
                        continue
                    min_len = min(len(x_values), len(observed_values), len(fitted_values))
                    x_values = x_values[:min_len]
                    observed_values = observed_values[:min_len]
                    fitted_values = fitted_values[:min_len]
                elif role == "residual_plot":
                    x_values = FigureStage.extract_named_series(columns, ["x", "time", "波数", "wavenumber", "index"])
                    y_values = FigureStage.extract_named_series(columns, ["residual", "res", "残差"])
                    if not x_values and y_values:
                        x_values = [float(i + 1) for i in range(len(y_values))]
                    if not x_values or not y_values:
                        missing.append({
                            "figure_request_id": req_id,
                            "required": True,
                            "satisfied": False,
                            "reason": "missing_residual_columns",
                        })
                        continue
                    min_len = min(len(x_values), len(y_values))
                    x_values = x_values[:min_len]
                    y_values = y_values[:min_len]
                else:
                    value_col = FigureStage.extract_named_series(columns, ["value", "metric", "score", "误差", "指标"])
                    if not value_col:
                        for _, values in columns.items():
                            if len(values) >= 2:
                                value_col = values
                                break
                    if not value_col:
                        missing.append({
                            "figure_request_id": req_id,
                            "required": True,
                            "satisfied": False,
                            "reason": "missing_comparison_columns",
                        })
                        continue
                    comparison_values = value_col[:16]
                    categories = [f"C{i + 1}" for i in range(len(comparison_values))]

            output_path = str(Path(work_dir) / "output" / f"{req_id}.png")
            title = (
                str(expected_content[0]).strip()
                if isinstance(expected_content, list) and expected_content and str(expected_content[0]).strip()
                else ("轨迹遮蔽过程" if chart_language == "Simplified Chinese" else "Trajectory Shielding")
            )

            if role == "trajectory_shielding":
                artifact = render_trajectory_shielding_2d(
                    points=list(zip(x_values, y_values)),
                    title=title,
                    output_path=output_path,
                    chart_language=chart_language,
                    figure_request_id=req_id,
                    subproblem_id=subproblem_id,
                    semantic_role="trajectory_shielding",
                    depicts=expected_content if isinstance(expected_content, list) else ["trajectory_shielding"],
                    linked_result_ids=[str(item.get("id") or "").strip() for item in (result_registry.get("verified_results", []) or []) if isinstance(item, dict) and str(item.get("id") or "").strip()][:12],
                    source_data=required_data_list or (matched_sources or ["result_registry.json"]),
                    work_dir=work_dir,
                )
            elif role == "distance_time_curve":
                artifact = render_distance_time_curve(
                    times=x_values,
                    distances=y_values,
                    title=title,
                    output_path=output_path,
                    chart_language=chart_language,
                    figure_request_id=req_id,
                    subproblem_id=subproblem_id,
                    semantic_role="distance_time_curve",
                    depicts=expected_content if isinstance(expected_content, list) else ["distance_time_curve"],
                    linked_result_ids=[str(item.get("id") or "").strip() for item in (result_registry.get("verified_results", []) or []) if isinstance(item, dict) and str(item.get("id") or "").strip()][:12],
                    source_data=required_data_list or (matched_sources or ["result_registry.json"]),
                    work_dir=work_dir,
                )
            elif role in {
                "geometry_schematic",
                "trajectory_overview",
                "optimization_variable_diagram",
                "physical_principle_schematic",
                "optical_path_diagram",
            }:
                artifact = render_geometry_schematic(
                    title=title,
                    output_path=output_path,
                    chart_language=chart_language,
                    figure_request_id=req_id,
                    subproblem_id=subproblem_id,
                    semantic_role=role,
                    depicts=expected_content if isinstance(expected_content, list) else [role],
                    linked_result_ids=[str(item.get("id") or "").strip() for item in (result_registry.get("verified_results", []) or []) if isinstance(item, dict) and str(item.get("id") or "").strip()][:12],
                    source_data=required_data_list or ["result_registry.json"],
                    model_objects=[str(item).strip() for item in model_objects if str(item).strip()] or ["model_core"],
                    formulas=[str(item).strip() for item in formulas if str(item).strip()],
                    source_problem_refs=[str(item).strip() for item in source_problem_refs if str(item).strip()] or ([subproblem_id] if subproblem_id else ["problem"]),
                    work_dir=work_dir,
                )
            elif role == "timeline_schematic":
                artifact = render_timeline_schematic(
                    title=title,
                    output_path=output_path,
                    chart_language=chart_language,
                    figure_request_id=req_id,
                    subproblem_id=subproblem_id,
                    semantic_role="timeline_schematic",
                    depicts=expected_content if isinstance(expected_content, list) else ["timeline_schematic"],
                    linked_result_ids=[str(item.get("id") or "").strip() for item in (result_registry.get("verified_results", []) or []) if isinstance(item, dict) and str(item.get("id") or "").strip()][:12],
                    source_data=required_data_list or ["result_registry.json"],
                    model_objects=[str(item).strip() for item in model_objects if str(item).strip()] or ["process_stage"],
                    formulas=[str(item).strip() for item in formulas if str(item).strip()],
                    source_problem_refs=[str(item).strip() for item in source_problem_refs if str(item).strip()] or ([subproblem_id] if subproblem_id else ["problem"]),
                    work_dir=work_dir,
                )
            elif role == "algorithm_flowchart":
                artifact = render_algorithm_flowchart(
                    title=title,
                    output_path=output_path,
                    chart_language=chart_language,
                    figure_request_id=req_id,
                    subproblem_id=subproblem_id,
                    semantic_role="algorithm_flowchart",
                    depicts=expected_content if isinstance(expected_content, list) else ["algorithm_flowchart"],
                    linked_result_ids=[str(item.get("id") or "").strip() for item in (result_registry.get("verified_results", []) or []) if isinstance(item, dict) and str(item.get("id") or "").strip()][:12],
                    source_data=required_data_list or ["result_registry.json"],
                    model_objects=[str(item).strip() for item in model_objects if str(item).strip()] or ["algorithm_pipeline"],
                    formulas=[str(item).strip() for item in formulas if str(item).strip()],
                    source_problem_refs=[str(item).strip() for item in source_problem_refs if str(item).strip()] or ([subproblem_id] if subproblem_id else ["problem"]),
                    work_dir=work_dir,
                )
            elif role == "spectrum_curve":
                artifact = render_spectrum_curve(
                    x_values=x_values,
                    y_values=y_values,
                    title=title,
                    output_path=output_path,
                    chart_language=chart_language,
                    figure_request_id=req_id,
                    subproblem_id=subproblem_id,
                    semantic_role="spectrum_curve",
                    depicts=expected_content if isinstance(expected_content, list) else ["spectrum_curve"],
                    linked_result_ids=[str(item.get("id") or "").strip() for item in (result_registry.get("verified_results", []) or []) if isinstance(item, dict) and str(item.get("id") or "").strip()][:12],
                    source_data=required_data_list or (matched_sources or ["result_registry.json"]),
                    work_dir=work_dir,
                )
            elif role == "fitting_curve":
                artifact = render_fitting_curve(
                    x_values=x_values,
                    observed_values=observed_values,
                    fitted_values=fitted_values,
                    title=title,
                    output_path=output_path,
                    chart_language=chart_language,
                    figure_request_id=req_id,
                    subproblem_id=subproblem_id,
                    semantic_role="fitting_curve",
                    depicts=expected_content if isinstance(expected_content, list) else ["fitting_curve"],
                    linked_result_ids=[str(item.get("id") or "").strip() for item in (result_registry.get("verified_results", []) or []) if isinstance(item, dict) and str(item.get("id") or "").strip()][:12],
                    source_data=required_data_list or (matched_sources or ["result_registry.json"]),
                    work_dir=work_dir,
                )
            elif role == "residual_plot":
                artifact = render_residual_plot(
                    x_values=x_values,
                    residual_values=y_values,
                    title=title,
                    output_path=output_path,
                    chart_language=chart_language,
                    figure_request_id=req_id,
                    subproblem_id=subproblem_id,
                    semantic_role="residual_plot",
                    depicts=expected_content if isinstance(expected_content, list) else ["residual_plot"],
                    linked_result_ids=[str(item.get("id") or "").strip() for item in (result_registry.get("verified_results", []) or []) if isinstance(item, dict) and str(item.get("id") or "").strip()][:12],
                    source_data=required_data_list or (matched_sources or ["result_registry.json"]),
                    work_dir=work_dir,
                )
            elif role == "comparison_bar":
                artifact = render_comparison_bar(
                    categories=categories,
                    values=comparison_values,
                    title=title,
                    output_path=output_path,
                    chart_language=chart_language,
                    figure_request_id=req_id,
                    subproblem_id=subproblem_id,
                    semantic_role="comparison_bar",
                    depicts=expected_content if isinstance(expected_content, list) else ["comparison_bar"],
                    linked_result_ids=[str(item.get("id") or "").strip() for item in (result_registry.get("verified_results", []) or []) if isinstance(item, dict) and str(item.get("id") or "").strip()][:12],
                    source_data=required_data_list or (matched_sources or ["result_registry.json"]),
                    work_dir=work_dir,
                )
            else:
                artifact = render_peak_valley_annotation(
                    x_values=x_values,
                    y_values=y_values,
                    title=title,
                    output_path=output_path,
                    chart_language=chart_language,
                    figure_request_id=req_id,
                    subproblem_id=subproblem_id,
                    semantic_role="peak_valley_annotation",
                    depicts=expected_content if isinstance(expected_content, list) else ["peak_valley_annotation"],
                    linked_result_ids=[str(item.get("id") or "").strip() for item in (result_registry.get("verified_results", []) or []) if isinstance(item, dict) and str(item.get("id") or "").strip()][:12],
                    source_data=required_data_list or (matched_sources or ["result_registry.json"]),
                    work_dir=work_dir,
                )

            if isinstance(artifact, dict):
                artifact["figure_kind"] = figure_kind or "result_figure"

            artifact_path = str(artifact.get("path") or "").strip() if isinstance(artifact, dict) else ""
            if artifact_path:
                created_images.append(artifact_path)
                satisfied.append({
                    "figure_request_id": req_id,
                    "required": True,
                    "satisfied": True,
                })
            else:
                missing.append({
                    "figure_request_id": req_id,
                    "required": True,
                    "satisfied": False,
                    "reason": "renderer_failed",
                })

        return {
            "created_images": list(dict.fromkeys(created_images)),
            "satisfied": satisfied,
            "missing": missing,
        }

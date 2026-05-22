"""FigurePlan builder for structured figure planning.

This module upgrades figure decision logic from keyword-only heuristics to a
multi-signal planner:
- problem/solve_spec structure signals
- domain template matching
- lightweight data-shape/file signals
- rule-based backfill when figure requests are missing
"""

from __future__ import annotations

import csv
import json
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from app.artifacts.protocols import (
    DEFAULT_FIGURE_CAPABILITIES,
    FigureRequestState,
    RESULT_OVERVIEW_ROLE,
    WorkflowIssue,
    canonical_caption_for_role,
    normalize_figure_role,
)
from app.artifacts.visualization_protocol import (
    VerifiedResult,
    normalize_verified_results,
    result_overview_request,
)
from app.utils.common_utils import save_json


PLACEHOLDER_RESULT_IDS = {
    "verified_series",
    "result_series",
    "placeholder_result",
    "todo_result",
}


def _norm_text(value: object) -> str:
    return str(value or "").strip()


def _join_texts(*parts: object) -> str:
    return "\n".join(_norm_text(p) for p in parts if _norm_text(p))


def _solve_spec_subproblems(solve_spec: dict[str, Any]) -> list[dict[str, Any]]:
    sub = solve_spec.get("subproblems", []) if isinstance(solve_spec, dict) else []
    return [item for item in sub if isinstance(item, dict)] if isinstance(sub, list) else []


@dataclass
class FigureNeedResult:
    requires_figures: bool
    requires_result_figures: bool
    requires_explanatory_figures: bool
    confidence: float
    reasons: list[str]
    detected_needs: list[str]
    domain_type: str


class FigureNeedDetector:
    """Multi-source figure-need detector."""

    OPTICAL_INTERFERENCE_RE = re.compile(
        r"外延层|衬底|干涉|多光束|光程差|波数|反射率|峰|谷|厚度公式|optical|interference|epitaxial|substrate",
        re.IGNORECASE,
    )

    EXPLANATORY_RE = re.compile(
        r"物理模型|几何|空间关系|轨迹|路径|时间过程|流程|推导|原理|"
        r"physical|geometry|trajectory|path|timeline|schematic|flow|derivation",
        re.IGNORECASE,
    )

    RESULT_RE = re.compile(
        r"测量数据|时间序列|优化|拟合|敏感性|比较|光谱|峰谷|反射率|"
        r"measurement|time\s*series|optimization|fit|sensitivity|comparison|spectrum|peak|valley|reflectance",
        re.IGNORECASE,
    )

    @classmethod
    def detect(
        cls,
        question: str,
        solve_spec: dict[str, Any],
        problem_facts: dict[str, Any] | None,
        work_dir: str,
    ) -> FigureNeedResult:
        reasons: list[str] = []
        detected_needs: list[str] = []

        combined = _join_texts(question, json.dumps(solve_spec or {}, ensure_ascii=False), json.dumps(problem_facts or {}, ensure_ascii=False))

        domain_type = "general"
        if cls.OPTICAL_INTERFERENCE_RE.search(combined):
            domain_type = "optical_thin_film_interference"
            reasons.append("domain_template_optical_thin_film_interference")
            detected_needs.extend([
                "optical_path_diagram",
                "spectrum_curve",
                "peak_valley_annotation",
                "algorithm_flowchart",
            ])

        # SolveSpec structure signals
        for sub in _solve_spec_subproblems(solve_spec):
            method = _norm_text(sub.get("method"))
            steps = sub.get("steps", [])
            expected_outputs = sub.get("expected_outputs", [])
            input_files = sub.get("input_files", [])

            if method:
                if cls.EXPLANATORY_RE.search(method):
                    reasons.append("solve_spec_method_requires_explanatory_figure")
                    detected_needs.append("physical_principle_schematic")
                if cls.RESULT_RE.search(method):
                    reasons.append("solve_spec_method_requires_result_figure")
                    detected_needs.append("result_data_plot")

            if isinstance(input_files, list) and input_files:
                reasons.append("solve_spec_has_input_files")
                detected_needs.append("result_data_plot")

            if isinstance(steps, list) and any("拟合" in _norm_text(step) or "fit" in _norm_text(step).lower() for step in steps):
                reasons.append("solve_spec_has_fitting_step")
                detected_needs.extend(["fitting_curve", "residual_plot"])

            if isinstance(expected_outputs, list) and any("厚度" in _norm_text(x) or "thickness" in _norm_text(x).lower() for x in expected_outputs):
                reasons.append("solve_spec_expected_thickness_output")
                detected_needs.append("algorithm_flowchart")

        # Data shape/file signals (lightweight)
        data_reasons, data_needs = cls._inspect_input_shapes(work_dir)
        reasons.extend(data_reasons)
        detected_needs.extend(data_needs)

        unique_needs = list(dict.fromkeys(detected_needs))
        unique_reasons = list(dict.fromkeys(reasons))

        requires_explanatory = any(
            n in unique_needs
            for n in {
                "physical_principle_schematic",
                "geometry_schematic",
                "optical_path_diagram",
                "timeline_schematic",
                "algorithm_flowchart",
            }
        )
        requires_result = any(
            n in unique_needs
            for n in {
                "spectrum_curve",
                "peak_valley_annotation",
                "fitting_curve",
                "residual_plot",
                "result_data_plot",
            }
        )

        requires_figures = requires_explanatory or requires_result
        confidence = min(0.99, 0.45 + 0.07 * len(unique_reasons) + (0.12 if domain_type != "general" else 0.0))
        if not requires_figures:
            confidence = 0.2

        return FigureNeedResult(
            requires_figures=requires_figures,
            requires_result_figures=requires_result,
            requires_explanatory_figures=requires_explanatory,
            confidence=round(confidence, 4),
            reasons=unique_reasons,
            detected_needs=unique_needs,
            domain_type=domain_type,
        )

    @classmethod
    def _inspect_input_shapes(cls, work_dir: str) -> tuple[list[str], list[str]]:
        reasons: list[str] = []
        needs: list[str] = []
        root = Path(work_dir)
        candidates: list[Path] = []
        for folder in (root / "data", root / "inputs", root):
            if not folder.exists() or not folder.is_dir():
                continue
            for f in folder.iterdir():
                if f.is_file() and f.suffix.lower() in {".csv", ".xlsx", ".xls"}:
                    candidates.append(f)

        if not candidates:
            return reasons, needs

        reasons.append("problem_has_measurement_data")

        for path in candidates[:8]:
            suffix = path.suffix.lower()
            name = path.name.lower()
            if suffix in {".xlsx", ".xls"}:
                reasons.append("tabular_excel_input_detected")
                needs.append("comparison_bar")
            if "spectrum" in name or "光谱" in name or "波数" in name:
                reasons.append("spectral_file_name_detected")
                needs.extend(["spectrum_curve", "peak_valley_annotation"])

            if suffix == ".csv":
                try:
                    with path.open("r", encoding="utf-8", newline="") as fp:
                        reader = csv.reader(fp)
                        header = next(reader, [])
                except Exception:
                    header = []
                header_text = " ".join(_norm_text(h).lower() for h in header)
                header_tokens = {
                    token
                    for h in header
                    for token in re.split(r"[^a-zA-Z0-9\u4e00-\u9fff]+", _norm_text(h).lower())
                    if token
                }
                if any(token in header_text for token in ["time", "时间", "t_"]):
                    reasons.append("time_series_column_detected")
                    needs.append("timeline_schematic")
                if any(token in header_text for token in ["wavenumber", "波数", "reflectance", "反射率"]):
                    reasons.append("spectrum_columns_detected")
                    needs.extend(["spectrum_curve", "peak_valley_annotation"])

                spatial_tokens = {"x", "y", "z", "x_coord", "y_coord", "z_coord", "坐标"}
                if len(header_tokens.intersection(spatial_tokens)) >= 2:
                    reasons.append("spatial_columns_detected")
                    needs.append("geometry_schematic")

        return list(dict.fromkeys(reasons)), list(dict.fromkeys(needs))


class FigurePlanBuilder:
    """Build and persist figure_plan.json as the source of truth."""

    EXPLANATORY_TEMPLATE_BY_NEED = {
        "physical_principle_schematic": ("physical_principle_schematic", True),
        "optical_path_diagram": ("optical_path_diagram", True),
        "geometry_schematic": ("geometry_schematic", True),
        "timeline_schematic": ("timeline_schematic", True),
        "algorithm_flowchart": ("algorithm_flowchart", True),
        "trajectory_overview": ("trajectory_overview", True),
        "optimization_variable_diagram": ("optimization_variable_diagram", False),
    }

    RESULT_TEMPLATE_BY_NEED = {
        "spectrum_curve": "spectrum_curve",
        "peak_valley_annotation": "peak_valley_annotation",
        "fitting_curve": "fitting_curve",
        "residual_plot": "residual_plot",
        "result_data_plot": "comparison_bar",
        "comparison_bar": "comparison_bar",
        "data_overview": RESULT_OVERVIEW_ROLE,
        "result_overview": RESULT_OVERVIEW_ROLE,
    }

    @classmethod
    def build(
        cls,
        question: str,
        solve_spec: dict[str, Any],
        problem_facts: dict[str, Any] | None,
        chart_language: str,
        work_dir: str,
    ) -> dict[str, Any]:
        need = FigureNeedDetector.detect(question, solve_spec, problem_facts, work_dir)
        requires_result_requested = bool(
            need.requires_result_figures
            or (isinstance(solve_spec, dict) and solve_spec.get("requires_result_figures") is True)
            or (isinstance(solve_spec, dict) and solve_spec.get("requires_result_figures_by_problem") is True)
        )
        requires_any_requested = bool(
            need.requires_figures
            or requires_result_requested
            or (isinstance(solve_spec, dict) and solve_spec.get("requires_figures") is True)
            or (isinstance(solve_spec, dict) and solve_spec.get("requires_figures_by_problem") is True)
        )
        requests = cls._seed_from_existing_requests(solve_spec, chart_language)
        available_data_files = cls._discover_input_files(work_dir)
        verified_result_ids = cls._load_verified_result_ids(work_dir)
        verified_results = normalize_verified_results(cls._load_result_registry(work_dir))

        # Rule-based backfill if detector says figures are needed but requests are empty/incomplete.
        if requires_any_requested:
            requests = cls._backfill_requests(
                requests=requests,
                solve_spec=solve_spec,
                chart_language=chart_language,
                domain_type=need.domain_type,
                detected_needs=need.detected_needs,
                available_data_files=available_data_files,
            )

        requests = cls._apply_capability_constraints(
            requests=requests,
            chart_language=chart_language,
            available_data_files=available_data_files,
            verified_result_ids=verified_result_ids,
            work_dir=work_dir,
        )

        if requires_any_requested and not any(
            isinstance(item, dict) and bool(item.get("required", True))
            for item in requests
        ):
            requests = cls._inject_fallback_result_overview_requests(
                requests=requests,
                solve_spec=solve_spec,
                chart_language=chart_language,
                available_data_files=available_data_files,
                verified_results=verified_results,
            )
            requests = cls._apply_capability_constraints(
                requests=requests,
                chart_language=chart_language,
                available_data_files=available_data_files,
                verified_result_ids=verified_result_ids,
                work_dir=work_dir,
            )

        if requires_result_requested and not any(
            isinstance(item, dict)
            and bool(item.get("required", True))
            and str(item.get("figure_kind") or "result_figure").strip().lower() == "result_figure"
            for item in requests
        ):
            requests = cls._inject_fallback_result_overview_requests(
                requests=requests,
                solve_spec=solve_spec,
                chart_language=chart_language,
                available_data_files=available_data_files,
                verified_results=verified_results,
            )
            requests = cls._apply_capability_constraints(
                requests=requests,
                chart_language=chart_language,
                available_data_files=available_data_files,
                verified_result_ids=verified_result_ids,
                work_dir=work_dir,
            )

        required_count = sum(1 for item in requests if isinstance(item, dict) and bool(item.get("required", True)))
        required_result_count = sum(
            1
            for item in requests
            if isinstance(item, dict)
            and bool(item.get("required", True))
            and str(item.get("figure_kind") or "result_figure").strip().lower() == "result_figure"
        )
        required_explanatory_count = sum(
            1
            for item in requests
            if isinstance(item, dict)
            and bool(item.get("required", True))
            and str(item.get("figure_kind") or "").strip().lower() == "explanatory_figure"
        )
        requires_result_figures = any(
            isinstance(item, dict)
            and bool(item.get("required", True))
            and str(item.get("figure_kind") or "result_figure").strip().lower() == "result_figure"
            for item in requests
        )
        requires_explanatory_figures = any(
            isinstance(item, dict)
            and bool(item.get("required", True))
            and str(item.get("figure_kind") or "").strip().lower() == "explanatory_figure"
            for item in requests
        )

        plan = {
            "requires_figures": bool(requires_any_requested),
            "requires_result_figures": bool(requires_result_requested),
            "requires_explanatory_figures": bool(need.requires_explanatory_figures),
            "requires_figures_by_problem": bool(requires_any_requested),
            "requires_result_figures_by_problem": bool(requires_result_requested),
            "requires_explanatory_figures_by_problem": bool(need.requires_explanatory_figures),
            "confidence": need.confidence,
            "reasons": need.reasons,
            "detected_needs": need.detected_needs,
            "domain_type": need.domain_type,
            "planning_basis": cls._planning_basis_from_reasons(need.reasons),
            "required_feasible_count": required_count,
            "required_feasible_result_count": required_result_count,
            "required_feasible_explanatory_count": required_explanatory_count,
            "required_count": required_count,
            "recommended_count": sum(1 for item in requests if isinstance(item, dict) and str(item.get("status") or "") == "recommended"),
            "blocked_count": sum(1 for item in requests if isinstance(item, dict) and str(item.get("status") or "") == "blocked"),
            "workflow_issues": [
                issue.to_dict()
                for issue in cls._issues_from_requests(requests)
            ],
            "figure_requests": requests,
        }

        return plan

    @classmethod
    def _discover_input_files(cls, work_dir: str) -> list[str]:
        root = Path(work_dir)
        discovered: list[str] = []
        for folder in (root / "inputs", root / "data", root):
            if not folder.exists() or not folder.is_dir():
                continue
            for path in folder.rglob("*"):
                if not path.is_file():
                    continue
                if path.suffix.lower() not in {".csv", ".xlsx", ".xls", ".json"}:
                    continue
                try:
                    rel = path.relative_to(root).as_posix()
                except Exception:
                    rel = path.name
                discovered.append(rel)
        return list(dict.fromkeys(discovered))

    @classmethod
    def _load_verified_result_ids(cls, work_dir: str) -> set[str]:
        path = Path(work_dir) / "result_registry.json"
        if not path.exists():
            return set()
        try:
            payload = json.loads(path.read_text(encoding="utf-8"))
        except Exception:
            return set()
        if not isinstance(payload, dict):
            return set()
        verified = payload.get("verified_results", [])
        if not isinstance(verified, list):
            return set()
        return {
            str(item.get("id") or "").strip()
            for item in verified
            if isinstance(item, dict) and str(item.get("id") or "").strip()
        }

    @classmethod
    def _load_result_registry(cls, work_dir: str) -> dict[str, Any]:
        path = Path(work_dir) / "result_registry.json"
        if not path.exists():
            return {}
        try:
            payload = json.loads(path.read_text(encoding="utf-8"))
        except Exception:
            return {}
        return payload if isinstance(payload, dict) else {}

    @classmethod
    def _apply_capability_constraints(
        cls,
        requests: list[dict[str, Any]],
        chart_language: str,
        available_data_files: list[str],
        verified_result_ids: set[str],
        work_dir: str,
    ) -> list[dict[str, Any]]:
        available_set = {str(item).strip().replace("\\", "/") for item in available_data_files if str(item).strip()}
        root = Path(work_dir)
        evaluated: list[dict[str, Any]] = []

        for raw in requests:
            if not isinstance(raw, dict):
                continue
            req = dict(raw)
            role = normalize_figure_role(req.get("semantic_role") or req.get("role"))
            kind = _norm_text(req.get("figure_kind") or "result_figure").lower()
            required = bool(req.get("required", True))
            req.setdefault("figure_kind", kind or "result_figure")
            req["semantic_role"] = role
            req.setdefault("role", role)

            depends_on = req.get("depends_on", {}) if isinstance(req.get("depends_on"), dict) else {}
            data_files = depends_on.get("data_files", []) if isinstance(depends_on.get("data_files"), list) else []
            data_files_list = [str(item).strip().replace("\\", "/") for item in data_files if str(item).strip()]

            linked_ids = req.get("linked_result_ids", [])
            if not isinstance(linked_ids, list):
                linked_ids = []
            if not linked_ids:
                linked_ids = depends_on.get("result_ids", []) if isinstance(depends_on.get("result_ids"), list) else []
            linked_ids_list = [str(item).strip() for item in linked_ids if str(item).strip()]
            placeholder_ids = {item for item in linked_ids_list if item.lower() in PLACEHOLDER_RESULT_IDS}
            linked_ids_list = [item for item in linked_ids_list if item.lower() not in PLACEHOLDER_RESULT_IDS]

            capability = DEFAULT_FIGURE_CAPABILITIES.get(role)
            if (
                kind == "result_figure"
                and bool(capability and capability.requires_verified_result)
                and not linked_ids_list
                and verified_result_ids
                and not placeholder_ids
            ):
                linked_ids_list = sorted(str(item).strip() for item in verified_result_ids if str(item).strip())[:12]
                req["auto_bound_verified_results"] = True

            req["linked_result_ids"] = linked_ids_list
            depends_on["result_ids"] = linked_ids_list
            depends_on["data_files"] = data_files_list
            req["depends_on"] = depends_on
            req["required_data"] = data_files_list

            blocked_reason = ""

            def _collect_header_tokens(files: list[str]) -> set[str]:
                tokens: set[str] = set()
                for rel in files:
                    path = Path(rel)
                    name = path.name.lower()
                    for token in re.split(r"[^a-zA-Z0-9\u4e00-\u9fff]+", name):
                        token = token.strip().lower()
                        if token:
                            tokens.add(token)
                    disk_path = root / rel
                    if not disk_path.exists() or not disk_path.is_file():
                        continue
                    try:
                        if disk_path.suffix.lower() == ".csv":
                            with disk_path.open("r", encoding="utf-8", newline="") as fp:
                                reader = csv.reader(fp)
                                header = next(reader, [])
                        elif disk_path.suffix.lower() in {".xlsx", ".xls"}:
                            from openpyxl import load_workbook  # type: ignore

                            wb = load_workbook(disk_path, read_only=True, data_only=True)
                            ws = wb.active
                            first_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), [])
                            header = list(first_row or [])
                        else:
                            header = []
                    except Exception:
                        header = []
                    for column in header:
                        for token in re.split(r"[^a-zA-Z0-9\u4e00-\u9fff]+", str(column or "").lower()):
                            token = token.strip().lower()
                            if token:
                                tokens.add(token)
                return tokens

            def _count_numeric_columns(files: list[str]) -> int:
                numeric_columns: set[str] = set()
                for rel in files:
                    disk_path = root / rel
                    if not disk_path.exists() or not disk_path.is_file():
                        continue
                    suffix = disk_path.suffix.lower()
                    try:
                        headers: list[str] = []
                        rows: list[list[object]] = []
                        if suffix == ".csv":
                            with disk_path.open("r", encoding="utf-8", newline="") as fp:
                                reader = csv.reader(fp)
                                headers = [str(item or "").strip() for item in next(reader, [])]
                                for _, row in zip(range(120), reader):
                                    rows.append(list(row))
                        elif suffix in {".xlsx", ".xls"}:
                            from openpyxl import load_workbook  # type: ignore

                            wb = load_workbook(disk_path, read_only=True, data_only=True)
                            ws = wb.active
                            it = ws.iter_rows(min_row=1, max_row=140, values_only=True)
                            header_row = next(it, [])
                            headers = [str(item or "").strip() for item in (header_row or [])]
                            for row in it:
                                rows.append(list(row or []))
                        else:
                            continue

                        for idx, name in enumerate(headers):
                            if not name:
                                continue
                            numeric_hits = 0
                            for row in rows:
                                val = row[idx] if idx < len(row) else ""
                                try:
                                    text = str(val or "").strip()
                                    if not text:
                                        continue
                                    float(text)
                                    numeric_hits += 1
                                except Exception:
                                    continue
                            if numeric_hits >= 2:
                                numeric_columns.add(name.lower())
                    except Exception:
                        continue
                return len(numeric_columns)

            request_context = "\n".join(
                [
                    _norm_text(req.get("purpose")),
                    _norm_text(req.get("chart_intent")),
                    " ".join(str(item) for item in req.get("required_columns", []) if str(item).strip()) if isinstance(req.get("required_columns"), list) else "",
                    " ".join(str(item) for item in req.get("expected_content", []) if str(item).strip()) if isinstance(req.get("expected_content"), list) else "",
                ]
            ).lower()
            header_tokens = _collect_header_tokens(data_files_list)

            if not capability:
                blocked_reason = f"unsupported_renderer:{role or 'unknown'}"
            elif chart_language not in set(capability.supports_languages):
                blocked_reason = f"unsupported_chart_language:{chart_language}"
            elif capability.domain_applicability and kind == "result_figure":
                applies = any(token.lower() in request_context for token in capability.domain_applicability)
                if not applies:
                    applies = any(token.lower() in header_tokens for token in capability.domain_applicability)
                if not applies:
                    blocked_reason = f"semantic_data_mismatch:{role}"
            if not blocked_reason and bool(capability and capability.requires_data) and kind == "result_figure":
                candidates = [item for item in data_files_list if item != "result_registry.json"]
                if not candidates:
                    blocked_reason = "missing_source_data"
                elif not any(item in available_set for item in candidates):
                    blocked_reason = "missing_source_data"
                elif role in {"fitting_curve", "residual_plot", "comparison_bar"}:
                    required_columns = [
                        str(item).strip().lower()
                        for item in (req.get("required_columns", []) if isinstance(req.get("required_columns"), list) else [])
                        if str(item).strip()
                    ]
                    if required_columns and not any(col in header_tokens for col in required_columns):
                        blocked_reason = f"required_columns_missing:{role}"
                elif role == RESULT_OVERVIEW_ROLE:
                    if _count_numeric_columns(candidates) < 2:
                        blocked_reason = f"insufficient_numeric_columns:{RESULT_OVERVIEW_ROLE}"
            if not blocked_reason and bool(capability and capability.requires_verified_result) and kind == "result_figure":
                if placeholder_ids:
                    blocked_reason = "placeholder_linked_result_ids"
                elif not linked_ids_list:
                    blocked_reason = "missing_verified_result_binding"
                elif not verified_result_ids:
                    blocked_reason = "pending_result_binding"
                elif verified_result_ids and any(item not in verified_result_ids for item in linked_ids_list):
                    blocked_reason = "linked_result_not_verified"

            if kind == "result_figure" and bool(capability and capability.requires_verified_result):
                if blocked_reason == "pending_result_binding":
                    req["result_binding_status"] = "pending_result_binding"
                elif not blocked_reason:
                    req["result_binding_status"] = "verified_result_bound"
                else:
                    req["result_binding_status"] = "binding_blocked"
            else:
                req["result_binding_status"] = "not_required"

            if blocked_reason:
                req["required"] = False
                req["must_include_in_paper"] = False
                req["recommended"] = True
                req["status"] = FigureRequestState.blocked.value
                req["blocked_reason"] = blocked_reason
                req["feasible"] = False
            else:
                req["feasible"] = True
                if required:
                    req["required"] = True
                    req["must_include_in_paper"] = bool(req.get("must_include_in_paper", True))
                    req["recommended"] = False
                    req["status"] = FigureRequestState.required.value
                    req.pop("blocked_reason", None)
                else:
                    req["required"] = False
                    req["must_include_in_paper"] = False
                    req["recommended"] = True
                    req["status"] = FigureRequestState.recommended.value
                    req.pop("blocked_reason", None)
            evaluated.append(req)

        return evaluated

    @classmethod
    def _issues_from_requests(cls, requests: list[dict[str, Any]]) -> list[WorkflowIssue]:
        issues: list[WorkflowIssue] = []
        for item in requests:
            if not isinstance(item, dict):
                continue
            if str(item.get("status") or "").strip() != FigureRequestState.blocked.value:
                continue
            request_id = str(item.get("id") or "").strip()
            reason = str(item.get("blocked_reason") or "").strip() or "blocked"
            issues.append(
                WorkflowIssue(
                    code="FIGURE_REQUEST_BLOCKED",
                    severity="warning",
                    stage="figure_plan",
                    reason=reason,
                    request_id=request_id,
                    action="Implement renderer, provide source data, or keep downgraded as recommended.",
                )
            )
        return issues

    @classmethod
    def save(cls, work_dir: str, figure_plan: dict[str, Any]) -> str:
        path = str(Path(work_dir) / "figure_plan.json")
        save_json(figure_plan, path)
        return path

    @classmethod
    def reevaluate_requests(
        cls,
        requests: list[dict[str, Any]],
        chart_language: str,
        work_dir: str,
    ) -> list[dict[str, Any]]:
        available_data_files = cls._discover_input_files(work_dir)
        verified_result_ids = cls._load_verified_result_ids(work_dir)
        verified_results = normalize_verified_results(cls._load_result_registry(work_dir))
        normalized = [dict(item) for item in requests if isinstance(item, dict)]
        had_required_request = any(
            isinstance(item, dict) and bool(item.get("required", True))
            for item in normalized
        )
        had_result_request = any(
            isinstance(item, dict)
            and str(item.get("figure_kind") or "result_figure").strip().lower() == "result_figure"
            for item in normalized
        )
        evaluated = cls._apply_capability_constraints(
            requests=normalized,
            chart_language=chart_language,
            available_data_files=available_data_files,
            verified_result_ids=verified_result_ids,
            work_dir=work_dir,
        )
        has_required_request = any(
            isinstance(item, dict) and bool(item.get("required", True))
            for item in evaluated
        )
        has_required_result_request = any(
            isinstance(item, dict)
            and bool(item.get("required", True))
            and str(item.get("figure_kind") or "result_figure").strip().lower() == "result_figure"
            for item in evaluated
        )
        if had_required_request and (not has_required_request or (had_result_request and not has_required_result_request)):
            pseudo_solve_spec = cls._solve_spec_from_requests(normalized)
            before_count = len(evaluated)
            evaluated = cls._inject_fallback_result_overview_requests(
                requests=evaluated,
                solve_spec=pseudo_solve_spec,
                chart_language=chart_language,
                available_data_files=available_data_files,
                verified_results=verified_results,
            )
            if len(evaluated) > before_count:
                evaluated = cls._apply_capability_constraints(
                    requests=evaluated,
                    chart_language=chart_language,
                    available_data_files=available_data_files,
                    verified_result_ids=verified_result_ids,
                    work_dir=work_dir,
                )
        return evaluated

    @classmethod
    def _solve_spec_from_requests(cls, requests: list[dict[str, Any]]) -> dict[str, Any]:
        subproblems: list[dict[str, Any]] = []
        seen: set[str] = set()
        for item in requests:
            if not isinstance(item, dict):
                continue
            sid = _norm_text(item.get("subproblem_id") or item.get("problem_section"))
            if not sid or sid in seen:
                continue
            seen.add(sid)
            data_files = []
            required_data = item.get("required_data", [])
            if isinstance(required_data, list):
                data_files = [str(src).strip() for src in required_data if str(src).strip()]
            subproblems.append(
                {
                    "id": sid,
                    "objective": _norm_text(item.get("purpose")) or sid,
                    "input_files": [src for src in data_files if src != "result_registry.json"],
                }
            )
        if not subproblems:
            subproblems = [{"id": "problem_1", "objective": "problem", "input_files": []}]
        return {"subproblems": subproblems}

    @classmethod
    def _planning_basis_from_reasons(cls, reasons: list[str]) -> list[str]:
        basis: list[str] = []
        mapping = {
            "problem_has_measurement_data": "problem_has_measurement_data",
            "solve_spec_has_input_files": "problem_has_measurement_data",
            "solve_spec_method_requires_explanatory_figure": "problem_has_physical_model",
            "solve_spec_method_requires_result_figure": "problem_has_algorithmic_process",
            "spectrum_columns_detected": "problem_has_measurement_data",
            "time_series_column_detected": "problem_has_algorithmic_process",
            "domain_template_optical_thin_film_interference": "problem_has_formula_derivation",
        }
        for reason in reasons:
            value = mapping.get(reason)
            if value:
                basis.append(value)
        return list(dict.fromkeys(basis))

    @classmethod
    def _seed_from_existing_requests(cls, solve_spec: dict[str, Any], chart_language: str) -> list[dict[str, Any]]:
        existing = solve_spec.get("figure_requests", []) if isinstance(solve_spec, dict) else []
        if not isinstance(existing, list):
            return []
        seeded: list[dict[str, Any]] = []
        for item in existing:
            if not isinstance(item, dict):
                continue
            payload = dict(item)
            payload.setdefault("chart_language", chart_language)
            payload.setdefault("language", chart_language)
            payload.setdefault("figure_kind", "result_figure")
            payload.setdefault("semantic_role", _norm_text(payload.get("role")) or "result_visualization")
            payload["semantic_role"] = normalize_figure_role(payload.get("semantic_role"))
            payload["role"] = payload["semantic_role"]
            payload.setdefault("must_include_in_paper", bool(payload.get("required", True)))
            payload.setdefault("depends_on", {
                "model_objects": [],
                "formulas": [],
                "data_files": payload.get("required_data", [] if not isinstance(payload.get("required_data"), list) else payload.get("required_data", [])),
                "result_ids": payload.get("linked_result_ids", []),
            })
            seeded.append(payload)
        return seeded

    @classmethod
    def _backfill_requests(
        cls,
        requests: list[dict[str, Any]],
        solve_spec: dict[str, Any],
        chart_language: str,
        domain_type: str,
        detected_needs: list[str],
        available_data_files: list[str],
    ) -> list[dict[str, Any]]:
        existing_ids = {
            _norm_text(item.get("id"))
            for item in requests
            if isinstance(item, dict) and _norm_text(item.get("id"))
        }

        subproblems = _solve_spec_subproblems(solve_spec)
        if not subproblems:
            subproblems = [{"id": "problem_1", "objective": _norm_text(solve_spec.get("summary")) or "problem"}]

        def _sub_id(index: int, sub: dict[str, Any]) -> str:
            return _norm_text(sub.get("id")) or f"problem_{index + 1}"

        def _resolve_data_files(sub: dict[str, Any], prefer_spectrum: bool = False) -> list[str]:
            input_files = sub.get("input_files", [])
            names = [str(item).strip() for item in input_files if str(item).strip()] if isinstance(input_files, list) else []
            matched: list[str] = []
            for name in names:
                name_base = Path(name).name.lower()
                for rel in available_data_files:
                    if Path(rel).name.lower() == name_base:
                        matched.append(rel)
            if not matched:
                matched = [rel for rel in available_data_files if rel.lower().endswith((".csv", ".xlsx", ".xls"))]
            if prefer_spectrum:
                spectral = [rel for rel in matched if any(token in rel.lower() for token in ["spectrum", "光谱", "波数", "reflectance"])]
                if spectral:
                    matched = spectral
            if "result_registry.json" not in matched:
                matched.append("result_registry.json")
            return list(dict.fromkeys(matched))

        def _mk_request(
            req_id: str,
            subproblem_id: str,
            figure_kind: str,
            semantic_role: str,
            purpose: str,
            required: bool,
            depends_on: dict[str, Any],
        ) -> dict[str, Any]:
            chart_intent = purpose
            required_columns: list[str] = []
            x_axis = ""
            y_axis = ""
            group_by = ""
            if semantic_role == "spectrum_curve":
                required_columns = ["wavenumber", "reflectance"]
                x_axis = "wavenumber"
                y_axis = "reflectance"
            elif semantic_role == "fitting_curve":
                required_columns = ["x", "observed", "fitted"]
                x_axis = "x"
                y_axis = "observed"
            elif semantic_role == "residual_plot":
                required_columns = ["x", "residual"]
                x_axis = "x"
                y_axis = "residual"
            elif semantic_role == "comparison_bar":
                required_columns = ["method", "metric", "value"]
                x_axis = "method"
                y_axis = "value"
                group_by = "metric"
            elif semantic_role == RESULT_OVERVIEW_ROLE:
                required_columns = []
                x_axis = "variables"
                y_axis = "numeric_value"
                group_by = ""
            canonical_caption = canonical_caption_for_role(semantic_role, chart_language)
            return {
                "id": req_id,
                "problem_section": subproblem_id,
                "subproblem_id": subproblem_id,
                "figure_kind": figure_kind,
                "semantic_role": semantic_role,
                "role": semantic_role,
                "canonical_caption": canonical_caption,
                "purpose": purpose,
                "chart_intent": chart_intent,
                "required": required,
                "must_include_in_paper": bool(required),
                "chart_language": chart_language,
                "language": chart_language,
                "required_columns": required_columns,
                "x_axis": x_axis,
                "y_axis": y_axis,
                "group_by": group_by,
                "view_type": "numeric_overview" if semantic_role == RESULT_OVERVIEW_ROLE else "",
                "depends_on": depends_on,
                "required_data": list(depends_on.get("data_files", []) or []),
                "linked_result_ids": list(depends_on.get("result_ids", []) or []),
                "expected_content": [purpose],
            }

        for i, sub in enumerate(subproblems):
            sid = _sub_id(i, sub)
            objective = _norm_text(sub.get("objective")) or sid

            # Domain-first template for optical interference
            if domain_type == "optical_thin_film_interference":
                optical_templates = [
                    ("optical_path_diagram", "explanatory_figure", "说明干涉光程差关系", True),
                    ("spectrum_curve", "result_figure", "展示反射率-波数光谱曲线", True),
                    ("peak_valley_annotation", "result_figure", "标注峰谷并支撑厚度估计", True),
                    ("algorithm_flowchart", "explanatory_figure", "说明厚度求解与复核流程", True),
                ]
                for role, kind, purpose, required in optical_templates:
                    req_id = f"fig_{sid}_{role}"
                    if req_id in existing_ids:
                        continue
                    existing_ids.add(req_id)
                    requests.append(
                        _mk_request(
                            req_id=req_id,
                            subproblem_id=sid,
                            figure_kind=kind,
                            semantic_role=role,
                            purpose=f"{objective}：{purpose}",
                            required=required,
                            depends_on={
                                "model_objects": ["epitaxial_layer", "substrate", "incident_beam", "reflected_beam"] if kind == "explanatory_figure" else [],
                                "formulas": ["2 n d cos(theta) = m lambda"] if role in {"optical_path_diagram", "peak_valley_annotation"} else [],
                                "data_files": ["result_registry.json"] if kind == "explanatory_figure" else _resolve_data_files(sub, prefer_spectrum=True),
                                "result_ids": [] if kind == "explanatory_figure" else ["thickness_estimate", "peak_wavenumbers"],
                            },
                        )
                    )

            # Generic detected-needs backfill
            for need in detected_needs:
                if need in cls.EXPLANATORY_TEMPLATE_BY_NEED:
                    role, required = cls.EXPLANATORY_TEMPLATE_BY_NEED[need]
                    req_id = f"fig_{sid}_{role}"
                    if req_id in existing_ids:
                        continue
                    existing_ids.add(req_id)
                    requests.append(
                        _mk_request(
                            req_id=req_id,
                            subproblem_id=sid,
                            figure_kind="explanatory_figure",
                            semantic_role=role,
                            purpose=f"{objective}：{role}",
                            required=required,
                            depends_on={
                                "model_objects": ["model_core"],
                                "formulas": [],
                                "data_files": ["result_registry.json"],
                                "result_ids": [],
                            },
                        )
                    )
                elif need in cls.RESULT_TEMPLATE_BY_NEED:
                    role = cls.RESULT_TEMPLATE_BY_NEED[need]
                    req_id = f"fig_{sid}_{role}"
                    if req_id in existing_ids:
                        continue
                    existing_ids.add(req_id)
                    requests.append(
                        _mk_request(
                            req_id=req_id,
                            subproblem_id=sid,
                            figure_kind="result_figure",
                            semantic_role=role,
                            purpose=f"{objective}：{role}",
                            required=True,
                            depends_on={
                                "model_objects": [],
                                "formulas": [],
                                "data_files": _resolve_data_files(sub, prefer_spectrum=(role in {"spectrum_curve", "peak_valley_annotation"})),
                                "result_ids": [],
                            },
                        )
                    )

        return requests

    @classmethod
    def _inject_fallback_data_overview_requests(
        cls,
        requests: list[dict[str, Any]],
        solve_spec: dict[str, Any],
        chart_language: str,
        available_data_files: list[str],
        verified_results: list[VerifiedResult] | None = None,
    ) -> list[dict[str, Any]]:
        """Backward-compatible wrapper for the canonical result overview fallback."""
        return cls._inject_fallback_result_overview_requests(
            requests=requests,
            solve_spec=solve_spec,
            chart_language=chart_language,
            available_data_files=available_data_files,
            verified_results=verified_results,
        )

    @classmethod
    def _inject_fallback_result_overview_requests(
        cls,
        requests: list[dict[str, Any]],
        solve_spec: dict[str, Any],
        chart_language: str,
        available_data_files: list[str],
        verified_results: list[VerifiedResult] | None = None,
    ) -> list[dict[str, Any]]:
        existing_ids = {
            _norm_text(item.get("id"))
            for item in requests
            if isinstance(item, dict) and _norm_text(item.get("id"))
        }
        existing_overview = next(
            (
                item for item in requests
                if isinstance(item, dict)
                and normalize_figure_role(item.get("semantic_role") or item.get("role")) == RESULT_OVERVIEW_ROLE
            ),
            None,
        )
        if existing_overview is not None:
            return cls._collapse_result_overview_requests(requests)

        subproblems = _solve_spec_subproblems(solve_spec)
        if not subproblems:
            subproblems = [{"id": "problem_1", "objective": "problem"}]

        def _resolve_data_files(sub: dict[str, Any]) -> list[str]:
            input_files = sub.get("input_files", [])
            names = [str(item).strip() for item in input_files if str(item).strip()] if isinstance(input_files, list) else []
            matched: list[str] = []
            for name in names:
                name_base = Path(name).name.lower()
                for rel in available_data_files:
                    if Path(rel).name.lower() == name_base:
                        matched.append(rel)
            if not matched:
                matched = [rel for rel in available_data_files if rel.lower().endswith((".csv", ".xlsx", ".xls"))]
            if "result_registry.json" not in matched:
                matched.append("result_registry.json")
            return list(dict.fromkeys(matched))

        for idx, sub in enumerate(subproblems, start=1):
            sid = _norm_text(sub.get("id")) or f"problem_{idx}"
            req_id = f"fig_{sid}_{RESULT_OVERVIEW_ROLE}"
            if req_id in existing_ids:
                continue
            objective = _norm_text(sub.get("objective")) or sid
            caption = canonical_caption_for_role(RESULT_OVERVIEW_ROLE, chart_language)
            data_files = _resolve_data_files(sub)
            matched_result = next(
                (
                    result for result in (verified_results or [])
                    if (result.section and result.section == sid)
                    or (not result.section and idx == 1)
                ),
                None,
            )
            result_ids_for_overview = [
                result.id
                for result in (verified_results or [])
                if result.id and (
                    (result.section and result.section == sid)
                    or (not result.section and idx == 1)
                )
            ]
            if not result_ids_for_overview and idx == 1:
                result_ids_for_overview = [
                    result.id for result in (verified_results or []) if result.id
                ][:12]
            if matched_result is not None or result_ids_for_overview:
                overview_result = matched_result or next(
                    (result for result in (verified_results or []) if result.id),
                    None,
                )
                if overview_result is None:
                    continue
                request = result_overview_request(
                    result=overview_result,
                    chart_language=chart_language,
                    data_files=data_files,
                )
                request["id"] = req_id
                request["problem_section"] = sid
                request["subproblem_id"] = sid
                request["purpose"] = f"{objective}: {caption}"
                request["depends_on"]["result_ids"] = result_ids_for_overview
                request["linked_result_ids"] = list(request["depends_on"]["result_ids"])
                requests.append(request)
                existing_ids.add(req_id)
                continue
            if verified_results:
                continue
            requests.append(
                {
                    "id": req_id,
                    "problem_section": sid,
                    "subproblem_id": sid,
                    "figure_kind": "result_figure",
                    "semantic_role": RESULT_OVERVIEW_ROLE,
                    "role": RESULT_OVERVIEW_ROLE,
                    "canonical_caption": caption,
                    "purpose": f"{objective}: {caption}",
                    "chart_intent": "numeric_overview",
                    "required": True,
                    "must_include_in_paper": True,
                    "chart_language": chart_language,
                    "language": chart_language,
                    "required_columns": [],
                    "x_axis": "variables",
                    "y_axis": "numeric_value",
                    "group_by": "",
                    "view_type": "numeric_overview",
                    "depends_on": {
                        "model_objects": [],
                        "formulas": [],
                        "data_files": data_files,
                        "result_ids": [],
                    },
                    "required_data": data_files,
                    "linked_result_ids": [],
                    "expected_content": [caption],
                }
            )
            existing_ids.add(req_id)
            break
        return cls._collapse_result_overview_requests(requests)

    @staticmethod
    def _collapse_result_overview_requests(requests: list[dict[str, Any]]) -> list[dict[str, Any]]:
        output: list[dict[str, Any]] = []
        overview: dict[str, Any] | None = None
        linked_ids: list[str] = []
        source_data: list[str] = []
        has_required_specific = any(
            isinstance(item, dict)
            and normalize_figure_role(item.get("semantic_role") or item.get("role")) != RESULT_OVERVIEW_ROLE
            and bool(item.get("required", True))
            and str(item.get("status") or "").strip() != "blocked"
            and not str(item.get("blocked_reason") or "").strip()
            for item in requests
        )
        for item in requests:
            if not isinstance(item, dict):
                continue
            role = normalize_figure_role(item.get("semantic_role") or item.get("role"))
            if role != RESULT_OVERVIEW_ROLE:
                output.append(item)
                continue
            if overview is None:
                overview = dict(item)
            for rid in item.get("linked_result_ids", []) or []:
                text = str(rid).strip()
                if text:
                    linked_ids.append(text)
            for src in item.get("required_data", []) or item.get("source_data", []) or []:
                text = str(src).strip()
                if text:
                    source_data.append(text)
        if overview is not None:
            overview["id"] = f"fig_{RESULT_OVERVIEW_ROLE}"
            overview["problem_section"] = "paper"
            overview["subproblem_id"] = "paper"
            overview["required"] = not has_required_specific
            overview["must_include_in_paper"] = not has_required_specific
            if linked_ids:
                overview["linked_result_ids"] = list(dict.fromkeys(linked_ids))[:12]
                overview.setdefault("depends_on", {})["result_ids"] = list(overview["linked_result_ids"])
            if source_data:
                overview["required_data"] = list(dict.fromkeys(source_data))
                overview.setdefault("depends_on", {})["data_files"] = list(overview["required_data"])
            output.append(overview)
        return output

import asyncio
import json
from pathlib import Path


class FakeDecomposer:
    async def decompose(self, *, question: str, data_files: list[str], task: dict):
        from app.guidance.contracts import ProblemSpec

        return ProblemSpec(
            task_summary=question,
            subproblems=[{"id": "problem_1", "objective": "fit branch flows"}],
            data_sources=[{"path": data_files[0]}],
        )


def test_decomposition_stage_saves_problem_spec_with_uploaded_files(tmp_path: Path) -> None:
    from app.guidance.stages import DecompositionStage

    data_dir = tmp_path / "data"
    data_dir.mkdir()
    (data_dir / "attachment.xlsx").write_bytes(b"xlsx")
    output = asyncio.run(DecompositionStage(decomposer=FakeDecomposer()).run(
        task_id="task-fit",
        task={"work_dir": str(tmp_path), "question": "Fit branch flows."},
        input_path=None,
        output_path=tmp_path / "problem_spec.json",
    ))

    payload = json.loads(output.read_text(encoding="utf-8"))
    assert payload["subproblems"][0]["id"] == "problem1"
    assert payload["data_sources"] == [{"path": "data/attachment.xlsx"}]


def test_decomposition_stage_marks_referenced_but_unuploaded_appendices_missing(tmp_path: Path) -> None:
    from app.guidance.contracts import ProblemSpec
    from app.guidance.stages import DecompositionStage

    class MisleadingDecomposer:
        async def decompose(self, *, question: str, data_files: list[str], task: dict):
            return ProblemSpec(
                task_summary=question,
                subproblems=[{"id": "problem1"}, {"id": "problem2"}],
                data_sources=[{
                    "path": "data/A-附件.xlsx",
                    "description": "该文件包含表1-4以及附录2和附录3。",
                }],
                locked_facts=[
                    "附录2提供刀盘载荷模型。",
                    "附录3提供滚刀磨损参数。",
                    "主附件提供预紧力观测。",
                ],
            )

    data_dir = tmp_path / "data"
    data_dir.mkdir()
    (data_dir / "A-附件.xlsx").write_bytes(b"xlsx")
    output = asyncio.run(DecompositionStage(decomposer=MisleadingDecomposer()).run(
        task_id="task-missing-appendices",
        task={
            "work_dir": str(tmp_path),
            "question": "问题2基于附录2计算载荷；问题3根据附录3分析磨损。",
        },
        input_path=None,
        output_path=tmp_path / "problem_spec.json",
    ))

    payload = json.loads(output.read_text(encoding="utf-8"))
    requirements = {item["reference"]: item for item in payload["source_requirements"]}
    assert requirements["附录2"]["status"] == "missing"
    assert requirements["附录3"]["status"] == "missing"
    assert all("附录2" not in fact and "附录3" not in fact for fact in payload["locked_facts"])
    assert any("附录2" in item and "未上传" in item for item in payload["uncertainties"])


def test_workbook_profile_normalizes_merged_groups_and_trailing_empty_columns(tmp_path: Path) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill

    from app.guidance.stages import _profile_workbook

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "标准实验数据"
    sheet.append(["滚刀直径", "破岩力"])
    sheet.merge_cells("A2:A4")
    sheet["A2"] = 17
    sheet["B2"] = 101.0
    sheet["B3"] = 102.0
    sheet["B4"] = 103.0
    sheet["C1"].fill = PatternFill(fill_type="solid", fgColor="FFFFFF")
    workbook_path = tmp_path / "merged-groups.xlsx"
    workbook.save(workbook_path)

    profile = _profile_workbook(workbook_path)["sheets"][0]

    assert profile["effective_max_column"] == 2
    assert profile["ignored_trailing_empty_columns"] == 1
    assert profile["columns"] == ["滚刀直径", "破岩力"]
    assert profile["merged_ranges"] == ["A2:A4"]
    assert profile["forward_fill_columns"] == ["滚刀直径"]
    assert profile["sample_rows"] == [[17, 101.0], [17, 102.0], [17, 103.0]]


def test_review_positive_value_accepts_consistent_with_explanation() -> None:
    from app.guidance.stages import _is_positive_review_value

    assert _is_positive_review_value("consistent; units are defined and aligned")

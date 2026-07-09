import asyncio
import json
from pathlib import Path

from openpyxl import Workbook


class WuyiProblemOneDecomposer:
    async def decompose(self, *, question: str, data_files: list[str], task: dict):
        from app.guidance.contracts import ProblemSpec

        return ProblemSpec(
            task_summary=question,
            subproblems=[{
                "id": "problem_1",
                "objective": "Infer two branch-flow functions from the observed main-road flow.",
                "problem_type": "constrained piecewise regression",
            }],
            data_sources=[{"path": data_files[0], "table": "Table 1"}],
            locked_facts=[
                "The main-road flow is the sum of branch 1 and branch 2.",
                "Branch 1 increases linearly; branch 2 increases then decreases linearly.",
            ],
            uncertainties=["The branch decomposition is not unique without an identification constraint."],
        )


class WuyiProblemOneModeler:
    async def model(self, *, problem_spec, task: dict, work_dir: Path):
        from app.guidance.contracts import ModelSpec

        return ModelSpec(
            model_id="wuyi-2025-a-problem-1",
            selected_method="continuous piecewise least squares with a symmetric-slope identification constraint",
            candidate_methods=[
                {
                    "name": "continuous piecewise least squares",
                    "selected": True,
                    "reason": "fits the stated piecewise trend under explicit constraints",
                },
                {"name": "unconstrained branch decomposition", "selected": False, "reason": "not identifiable"},
            ],
            assumptions=[
                "The absolute increase and decrease slopes of branch 2 are equal; this explicit constraint identifies the branch split."
            ],
            symbols=[
                {"symbol": "t", "meaning": "two-minute sample index"},
                {"symbol": "q_1", "meaning": "branch 1 flow"},
                {"symbol": "q_2", "meaning": "branch 2 flow"},
            ],
            subproblem_models=[{
                "subproblem_id": "problem_1",
                "objective": "Infer branch 1 and branch 2 flow functions.",
                "selected_method": "continuous piecewise least squares",
                "rationale": "The symmetric-slope constraint identifies the branch split.",
                "equations": [
                    "q_main(t) = q_1(t) + q_2(t)",
                    "q_1(t) = b_1 + a_1 t",
                    "q_2(t) has equal-magnitude slopes around t_break",
                ],
                "parameters": [
                    {
                        "parameter_id": "breakpoint",
                        "symbol": "t_break",
                        "meaning": "branch 2 turning point",
                        "unit": "two-minute sample",
                        "source_type": "estimated",
                        "source_detail": "data/attachment.xlsx Table 1",
                        "estimation_method": "least-squares breakpoint scan",
                        "constraints": ["2 <= t_break <= max(t)-1"],
                    },
                    {
                        "parameter_id": "branch1_slope",
                        "symbol": "a_1",
                        "meaning": "branch 1 slope",
                        "unit": "flow/sample",
                        "source_type": "estimated",
                        "source_detail": "data/attachment.xlsx Table 1",
                        "estimation_method": "aggregate segment-slope decomposition",
                        "constraints": ["a_1 > 0"],
                    },
                    {
                        "parameter_id": "branch2_slope",
                        "symbol": "a_2",
                        "meaning": "absolute branch 2 slope",
                        "unit": "flow/sample",
                        "source_type": "estimated",
                        "source_detail": "data/attachment.xlsx Table 1",
                        "estimation_method": "symmetric-slope identification",
                        "constraints": ["a_2 > 0"],
                    },
                ],
                "constraints": ["q_1(t) >= 0", "q_2(t) >= 0", "q_2 is continuous at t_break"],
                "solve_steps": [
                    "Read Table 1 from the XLSX attachment.",
                    "Scan breakpoints and minimize squared residuals.",
                    "Apply the symmetric-slope constraint to recover both branches.",
                    "Recompose main-road flow and calculate residual metrics.",
                ],
                "expected_results": ["two branch functions", "RMSE", "parameter stability"],
            }],
            figure_requests=[{
                "path": "figures/problem1_fit.png",
                "role": "result_fit",
                "linked_result_ids": ["result_problem1_piecewise_fit"],
                "source_data": ["data/attachment.xlsx"],
            }],
        )


class WuyiProblemOneReviewer:
    async def review(self, *, problem_spec, model_spec, task: dict, work_dir: Path):
        from app.guidance.contracts import ModelReview

        return ModelReview(
            model_id=model_spec.model_id,
            status="approved",
            identifiability="conditional_on_symmetric_branch2_slopes",
            dimensional_consistency="passed",
            data_sufficiency="passed_for_piecewise_fit",
            computational_feasibility="passed",
        )


class WuyiRevisionModeler(WuyiProblemOneModeler):
    async def model(self, *, problem_spec, task: dict, work_dir: Path):
        model = await super().model(
            problem_spec=problem_spec,
            task=task,
            work_dir=work_dir,
        )
        return model.model_copy(
            update={
                "selected_method": "unconstrained piecewise least squares",
                "assumptions": [],
            }
        )

    async def revise(
        self,
        *,
        problem_spec,
        previous_model,
        review,
        task: dict,
        work_dir: Path,
    ):
        from app.guidance.contracts import RevisedModelSpec

        revised = await WuyiProblemOneModeler().model(
            problem_spec=problem_spec,
            task=task,
            work_dir=work_dir,
        )
        return RevisedModelSpec.model_validate(
            {
                **revised.model_dump(mode="json"),
                "review_resolutions": [{
                    "requirement": review.required_revisions[0],
                    "resolution": "Added the symmetric-slope identification constraint.",
                    "affected_subproblem_ids": ["problem_1"],
                    "changed_fields": ["assumptions", "subproblem_models.0.constraints"],
                }],
            }
        )


class WuyiRevisionReviewer:
    def __init__(self) -> None:
        self.calls = 0

    async def review(self, *, problem_spec, model_spec, task: dict, work_dir: Path):
        from app.guidance.contracts import ModelReview

        self.calls += 1
        identified = any("identifies the branch split" in item for item in model_spec.assumptions)
        if identified:
            return ModelReview(
                model_id=model_spec.model_id,
                status="approved",
                identifiability="conditional_on_symmetric_branch2_slopes",
                dimensional_consistency="passed",
                data_sufficiency="passed_for_piecewise_fit",
                computational_feasibility="passed",
            )
        return ModelReview(
            model_id=model_spec.model_id,
            status="revision_required",
            identifiability="failed",
            dimensional_consistency="passed",
            data_sufficiency="passed_for_piecewise_fit",
            computational_feasibility="passed",
            blocking_issues=["The branch split is not uniquely identifiable."],
            required_revisions=["Add an independently justified identification constraint."],
        )


class WuyiProblemOneProgramGenerator:
    async def generate(self, *, approved_model, task: dict, work_dir: Path, **kwargs):
        from app.guidance.contracts import GeneratedProgram

        return GeneratedProgram(
            source_code=WUYI_PROBLEM_ONE_SOLVER,
            declared_outputs=approved_model.required_outputs,
        )


WUYI_PROBLEM_ONE_SOLVER = r'''
from pathlib import Path
import json
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt

source = Path("data/attachment.xlsx")
source_ref = source.as_posix()
frame = pd.read_excel(source, sheet_name=0)
t = frame.iloc[:, 1].astype(float).to_numpy()
y = frame.iloc[:, 2].astype(float).to_numpy()

def fit_piecewise(x, values):
    best_fit = None
    for breakpoint in range(2, int(np.max(x)) - 1):
        design = np.column_stack([np.ones_like(x), x, np.maximum(0.0, x - breakpoint)])
        coefficients, _, _, _ = np.linalg.lstsq(design, values, rcond=None)
        fitted = design @ coefficients
        sse = float(np.sum((values - fitted) ** 2))
        if best_fit is None or sse < best_fit[0]:
            best_fit = (sse, breakpoint, coefficients, fitted)
    return best_fit

sse, t_break, coefficients, fitted = fit_piecewise(t, y)
intercept = float(coefficients[0])
slope_before = float(coefficients[1])
slope_after = float(coefficients[1] + coefficients[2])
branch1_slope = (slope_before + slope_after) / 2.0
branch2_slope = (slope_before - slope_after) / 2.0
branch1 = intercept + branch1_slope * t
branch2 = np.where(
    t <= t_break,
    branch2_slope * t,
    branch2_slope * (2.0 * t_break - t),
)
recomposed = branch1 + branch2
residual = y - recomposed
rmse = float(np.sqrt(np.mean(residual ** 2)))
max_abs_residual = float(np.max(np.abs(residual)))
base_parameters = np.array([t_break, intercept, branch1_slope, branch2_slope], dtype=float)
relative_changes = []
for indices in (np.arange(1, len(t)), np.arange(0, len(t) - 1)):
    _, varied_break, varied_coefficients, _ = fit_piecewise(t[indices], y[indices])
    varied_before = float(varied_coefficients[1])
    varied_after = float(varied_coefficients[1] + varied_coefficients[2])
    varied_parameters = np.array([
        varied_break,
        float(varied_coefficients[0]),
        (varied_before + varied_after) / 2.0,
        (varied_before - varied_after) / 2.0,
    ])
    relative_changes.extend(
        (np.abs(varied_parameters - base_parameters) / np.maximum(np.abs(base_parameters), 1e-12)).tolist()
    )
max_relative_change = float(max(relative_changes))

parameters = [
    {"id": "param_t_break", "symbol": "t_break", "name": "breakpoint", "value": int(t_break), "unit": "sample index", "source_type": "estimated_from_data", "source_ref": source_ref, "trust_status": "estimated"},
    {"id": "param_branch1_intercept", "symbol": "b_1", "name": "branch 1 intercept", "value": intercept, "unit": "flow", "source_type": "estimated_from_data", "source_ref": source_ref, "trust_status": "estimated"},
    {"id": "param_branch1_slope", "symbol": "a_1", "name": "branch 1 slope", "value": branch1_slope, "unit": "flow/sample", "source_type": "estimated_from_data", "source_ref": source_ref, "trust_status": "estimated"},
    {"id": "param_branch2_slope", "symbol": "a_2", "name": "branch 2 slope magnitude", "value": branch2_slope, "unit": "flow/sample", "source_type": "estimated_from_data", "source_ref": source_ref, "trust_status": "estimated"},
]
result = {
    "id": "result_problem1_piecewise_fit",
    "result_type": "regression",
    "claim_text": "The constrained branch-flow model recomposes the observed main-road flow.",
    "metrics": {"rmse": rmse, "max_abs_residual": max_abs_residual},
    "source_data": [source_ref],
}
summary = {
    "verified_count": 1,
    "blocked_count": 0,
    "coverage_status": "verified",
    "reader_decision_guide": [
        "Observed variable is the main-road aggregate flow; branch flows are unknown.",
        "Time context uses sample index t from the worksheet rows.",
        "The branch split is conditional on the explicit identifiability constraint and must not be written as an unconstrained unique recovery.",
    ],
    "method_route_comparison": [
        {
            "route": "constrained piecewise least squares",
            "decision": "recommended",
            "reason": "matches the stated trend and exposes residuals",
            "boundary": "conditional branch-flow estimate",
            "evidence": [result["id"]],
        },
        {
            "route": "unconstrained unique branch recovery",
            "decision": "rejected",
            "reason": "aggregate observations alone leave branch-flow degrees of freedom",
            "boundary": "requires branch sensors or additional prior constraints",
            "evidence": [result["id"]],
        }
    ],
    "identifiability_analysis": {
        "observed": "main-road aggregate flow",
        "unknowns": "branch 1 and branch 2 flows",
        "rank_check": "conditional on explicit split constraint",
        "conclusion": "representative estimate, not unconstrained unique recovery",
    },
    "claim_registry": [
        {
            "claim": result["claim_text"],
            "status": "evidence_verified",
            "evidence": [result["id"]],
        }
    ],
    "final_answer_tables": [
        {
            "table": "problem 1 branch-flow functions",
            "status": "conditional_estimate",
            "result_id": result["id"],
        }
    ],
}

Path("parameter_registry.json").write_text(json.dumps({"parameters": parameters}, ensure_ascii=False, indent=2), encoding="utf-8")
Path("result_registry.json").write_text(json.dumps({"verified_results": [result], "blocked_results": [], "summary": summary}, ensure_ascii=False, indent=2), encoding="utf-8")
solver_results = {"evidence": [{
    "kind": "regression",
    "subproblem_id": "problem_1",
    "result_id": result["id"],
    "observed": y.tolist(),
    "predicted": recomposed.tolist(),
    "residuals": residual.tolist(),
    "constraint_values": np.concatenate([branch1, branch2]).tolist(),
    "parameter_stability": {
        "method": "endpoint leave-one-out refit",
        "max_relative_change": max_relative_change,
        "threshold": 0.10,
    },
}]}
Path("solver_results.json").write_text(json.dumps(solver_results, ensure_ascii=False), encoding="utf-8")
pd.DataFrame(parameters).to_csv("parameter_table.csv", index=False)

Path("figures").mkdir(exist_ok=True)
fig, ax = plt.subplots(figsize=(8, 5))
ax.plot(t, y, "o", markersize=3, label="Observed main road")
ax.plot(t, recomposed, "-", label="Recomposed flow")
ax.plot(t, branch1, "--", label="Branch 1")
ax.plot(t, branch2, "--", label="Branch 2")
ax.set_xlabel("Sample index")
ax.set_ylabel("Traffic flow")
ax.legend()
fig.tight_layout()
fig.savefig("figures/problem1_fit.png", dpi=160)
plt.close(fig)
figure_registry = {"figures": [{"path": "figures/problem1_fit.png", "role": "result_fit", "source_data": [source_ref], "linked_result_ids": [result["id"]]}]}
Path("figure_registry.json").write_text(json.dumps(figure_registry, ensure_ascii=False, indent=2), encoding="utf-8")
print("complete solver finished")
'''.strip()


def _write_problem_one_xlsx(path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Table 1"
    sheet.append(["Moment", "Time t", "Main road 3 flow"])
    for t in range(60):
        flow = 7.0 + 1.5 * t if t <= 30 else 67.0 - 0.5 * t
        sheet.append([f"sample-{t}", t, flow])
    workbook.save(path)


def test_wuyi_problem_one_runs_real_xlsx_jupyter_and_png_end_to_end(tmp_path: Path) -> None:
    from app.guidance.execution import LocalProgramExecutor
    from app.guidance.pipeline import GuidancePipeline
    from app.guidance.stages import (
        CalculationStage,
        DecompositionStage,
        GuidanceStage,
        ModelingStage,
        ModelReviewStage,
        ResultVerificationStage,
    )

    data_dir = tmp_path / "data"
    data_dir.mkdir()
    _write_problem_one_xlsx(data_dir / "attachment.xlsx")
    reviewer = WuyiRevisionReviewer()
    pipeline = GuidancePipeline(
        decomposition_stage=DecompositionStage(decomposer=WuyiProblemOneDecomposer()),
        modeling_stage=ModelingStage(modeler=WuyiRevisionModeler()),
        model_review_stage=ModelReviewStage(reviewer=reviewer),
        calculation_stage=CalculationStage(
            program_generator=WuyiProblemOneProgramGenerator(),
            executor=LocalProgramExecutor(),
        ),
        result_verification_stage=ResultVerificationStage(),
        guidance_stage=GuidanceStage(),
    )
    task = {
        "task_id": "wuyi-problem-one",
        "question": "Infer branch 1 and branch 2 traffic-flow functions for Problem 1.",
        "work_dir": str(tmp_path),
    }

    async def collect():
        return [message async for message in pipeline.run("wuyi-problem-one", task)]

    messages = asyncio.run(collect())

    parameters = json.loads((tmp_path / "parameter_registry.json").read_text(encoding="utf-8"))
    parameter_values = {item["id"]: item["value"] for item in parameters["parameters"]}
    results = json.loads((tmp_path / "result_registry.json").read_text(encoding="utf-8"))
    verification = json.loads((tmp_path / "verification_report.json").read_text(encoding="utf-8"))
    guidance = (tmp_path / "guidance.md").read_text(encoding="utf-8")
    audit = json.loads((tmp_path / "guidance_audit_report.json").read_text(encoding="utf-8"))

    assert parameter_values["param_t_break"] == 30
    assert reviewer.calls == 2
    assert abs(parameter_values["param_branch1_slope"] - 0.5) < 1e-10
    assert abs(parameter_values["param_branch2_slope"] - 1.0) < 1e-10
    assert results["verified_results"][0]["metrics"]["rmse"] < 1e-10
    assert verification["status"] == "verified"
    assert verification["sensitivity_checks"][0]["status"] == "passed"
    assert verification["artifact_refs"]["figures"][0]["linked_result_ids"] == [
        "result_problem1_piecewise_fit"
    ]
    assert (tmp_path / "execution_manifest.json").is_file()
    assert json.loads((tmp_path / "execution_manifest.json").read_text(encoding="utf-8"))["execution_count"] == 1
    assert (tmp_path / "notebook.ipynb").is_file()
    assert (tmp_path / "figures" / "problem1_fit.png").stat().st_size > 1000
    assert "param_branch1_slope" in guidance
    assert "result_problem1_piecewise_fit" in guidance
    assert "data/attachment.xlsx" in guidance
    assert "可辨识" in guidance
    assert audit["status"] == "PASS"
    assert messages[-1]["data"]["status"] == "completed"


class ProductMixOptimizationDecomposer:
    async def decompose(self, *, question: str, data_files: list[str], task: dict):
        from app.guidance.contracts import ProblemSpec

        return ProblemSpec(
            task_summary=question,
            subproblems=[{
                "id": "problem_1",
                "objective": "Choose product quantities that maximize total profit under capacity and demand constraints.",
                "problem_type": "linear optimization",
            }],
            data_sources=[{"path": data_files[0], "table": "products and capacities"}],
            locked_facts=[
                "Each product consumes two limited materials.",
                "Production cannot exceed demand limits.",
            ],
            uncertainties=["The fixture treats production quantities as continuous for first-pass guidance."],
        )


class ProductMixOptimizationModeler:
    async def model(self, *, problem_spec, task: dict, work_dir: Path):
        from app.guidance.contracts import ModelSpec

        return ModelSpec(
            model_id="product-mix-optimization-fixture",
            selected_method="two-variable linear programming over feasible vertices",
            candidate_methods=[
                {
                    "name": "linear programming",
                    "selected": True,
                    "reason": "the objective and all constraints are linear in production quantities",
                },
                {
                    "name": "greedy profit ranking",
                    "selected": False,
                    "reason": "greedy ranking can violate coupled resource constraints",
                },
            ],
            assumptions=[
                "Production quantities are continuous in the guidance fixture.",
                "Unit profit and material usage are known from the uploaded workbook.",
            ],
            symbols=[
                {"symbol": "x_A", "meaning": "quantity of product A"},
                {"symbol": "x_B", "meaning": "quantity of product B"},
                {"symbol": "P", "meaning": "total profit"},
            ],
            subproblem_models=[{
                "subproblem_id": "problem_1",
                "objective": "Maximize total profit under resource and demand limits.",
                "selected_method": "linear programming",
                "rationale": "The feasible region is a convex polygon, so an optimum occurs at a vertex.",
                "equations": [
                    "maximize P = 30 x_A + 40 x_B",
                    "2 x_A + x_B <= 100",
                    "x_A + 3 x_B <= 90",
                ],
                "parameters": [
                    {
                        "parameter_id": "profit_A",
                        "symbol": "p_A",
                        "meaning": "unit profit of product A",
                        "unit": "currency/unit",
                        "source_type": "data",
                        "source_detail": "data/optimization.xlsx products sheet",
                        "estimation_method": "read from workbook",
                        "constraints": ["p_A > 0"],
                    },
                    {
                        "parameter_id": "profit_B",
                        "symbol": "p_B",
                        "meaning": "unit profit of product B",
                        "unit": "currency/unit",
                        "source_type": "data",
                        "source_detail": "data/optimization.xlsx products sheet",
                        "estimation_method": "read from workbook",
                        "constraints": ["p_B > 0"],
                    },
                    {
                        "parameter_id": "capacity_material_a",
                        "symbol": "C_A",
                        "meaning": "capacity of material A",
                        "unit": "material units",
                        "source_type": "data",
                        "source_detail": "data/optimization.xlsx capacities sheet",
                        "estimation_method": "read from workbook",
                        "constraints": ["C_A >= 0"],
                    },
                ],
                "constraints": [
                    "x_A >= 0",
                    "x_B >= 0",
                    "x_A <= 40",
                    "x_B <= 50",
                    "2 x_A + x_B <= 100",
                    "x_A + 3 x_B <= 90",
                ],
                "solve_steps": [
                    "Read product coefficients and capacities from the XLSX attachment.",
                    "Enumerate all feasible vertices of the two-variable linear program.",
                    "Select the vertex with the largest total profit.",
                    "Register the objective value, resource usage, constraints and sensitivity evidence.",
                ],
                "expected_results": ["optimal product quantities", "total profit", "binding constraints"],
            }],
            figure_requests=[{
                "path": "figures/product_mix_feasible_region.png",
                "role": "verification_diagnostic",
                "linked_result_ids": ["result_product_mix_optimum"],
                "source_data": ["data/optimization.xlsx"],
            }],
        )


class ProductMixOptimizationReviewer:
    async def review(self, *, problem_spec, model_spec, task: dict, work_dir: Path):
        from app.guidance.contracts import ModelReview

        return ModelReview(
            model_id=model_spec.model_id,
            status="approved",
            identifiability="not_applicable_for_deterministic_optimization",
            dimensional_consistency="passed",
            data_sufficiency="passed_for_two_variable_fixture",
            computational_feasibility="passed",
        )


class ProductMixOptimizationProgramGenerator:
    async def generate(self, *, approved_model, task: dict, work_dir: Path, **kwargs):
        from app.guidance.contracts import GeneratedProgram

        return GeneratedProgram(
            source_code=PRODUCT_MIX_OPTIMIZATION_SOLVER,
            declared_outputs=approved_model.required_outputs,
        )


PRODUCT_MIX_OPTIMIZATION_SOLVER = r'''
from pathlib import Path
import itertools
import json
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt

source = Path("data/optimization.xlsx")
source_ref = source.as_posix()
products = pd.read_excel(source, sheet_name="products")
capacities = pd.read_excel(source, sheet_name="capacities")

profit = products["unit_profit"].astype(float).to_numpy()
resource_matrix = products[["material_a", "material_b"]].astype(float).to_numpy().T
demand = products["max_demand"].astype(float).to_numpy()
capacity = capacities["capacity"].astype(float).to_numpy()

constraints = [
    (np.array([resource_matrix[0, 0], resource_matrix[0, 1]], dtype=float), capacity[0], "material_a_capacity"),
    (np.array([resource_matrix[1, 0], resource_matrix[1, 1]], dtype=float), capacity[1], "material_b_capacity"),
    (np.array([1.0, 0.0]), demand[0], "product_a_demand"),
    (np.array([0.0, 1.0]), demand[1], "product_b_demand"),
    (np.array([-1.0, 0.0]), 0.0, "product_a_nonnegative"),
    (np.array([0.0, -1.0]), 0.0, "product_b_nonnegative"),
]

candidates = []
for left, right in itertools.combinations(constraints, 2):
    matrix = np.vstack([left[0], right[0]])
    if abs(np.linalg.det(matrix)) < 1e-10:
        continue
    point = np.linalg.solve(matrix, np.array([left[1], right[1]], dtype=float))
    feasible = all(float(a @ point) <= b + 1e-8 for a, b, _ in constraints)
    if feasible:
        candidates.append(point)
candidates.append(np.array([0.0, 0.0]))

best = max(candidates, key=lambda point: float(profit @ point))
usage = resource_matrix @ best
total_profit = float(profit @ best)
objective_value = -total_profit
baseline_objective = 0.0

parameters = [
    {"id": "param_profit_a", "symbol": "p_A", "name": "unit profit of product A", "value": float(profit[0]), "unit": "currency/unit", "source_type": "read_from_data", "source_ref": source_ref, "trust_status": "given"},
    {"id": "param_profit_b", "symbol": "p_B", "name": "unit profit of product B", "value": float(profit[1]), "unit": "currency/unit", "source_type": "read_from_data", "source_ref": source_ref, "trust_status": "given"},
    {"id": "param_material_a_per_product_a", "symbol": "a_A", "name": "material A usage per product A", "value": float(resource_matrix[0, 0]), "unit": "material units/unit", "source_type": "read_from_data", "source_ref": source_ref, "trust_status": "given"},
    {"id": "param_material_a_per_product_b", "symbol": "a_B", "name": "material A usage per product B", "value": float(resource_matrix[0, 1]), "unit": "material units/unit", "source_type": "read_from_data", "source_ref": source_ref, "trust_status": "given"},
    {"id": "param_material_b_per_product_a", "symbol": "b_A", "name": "material B usage per product A", "value": float(resource_matrix[1, 0]), "unit": "material units/unit", "source_type": "read_from_data", "source_ref": source_ref, "trust_status": "given"},
    {"id": "param_material_b_per_product_b", "symbol": "b_B", "name": "material B usage per product B", "value": float(resource_matrix[1, 1]), "unit": "material units/unit", "source_type": "read_from_data", "source_ref": source_ref, "trust_status": "given"},
    {"id": "param_material_a_capacity", "symbol": "C_A", "name": "material A capacity", "value": float(capacity[0]), "unit": "material units", "source_type": "read_from_data", "source_ref": source_ref, "trust_status": "given"},
    {"id": "param_material_b_capacity", "symbol": "C_B", "name": "material B capacity", "value": float(capacity[1]), "unit": "material units", "source_type": "read_from_data", "source_ref": source_ref, "trust_status": "given"},
    {"id": "param_product_a_demand", "symbol": "D_A", "name": "maximum demand for product A", "value": float(demand[0]), "unit": "units", "source_type": "read_from_data", "source_ref": source_ref, "trust_status": "given"},
    {"id": "param_product_b_demand", "symbol": "D_B", "name": "maximum demand for product B", "value": float(demand[1]), "unit": "units", "source_type": "read_from_data", "source_ref": source_ref, "trust_status": "given"},
    {"id": "param_product_a_quantity", "symbol": "x_A", "name": "optimal quantity of product A", "value": float(best[0]), "unit": "units", "source_type": "computed_from_data", "source_ref": source_ref, "trust_status": "verified"},
    {"id": "param_product_b_quantity", "symbol": "x_B", "name": "optimal quantity of product B", "value": float(best[1]), "unit": "units", "source_type": "computed_from_data", "source_ref": source_ref, "trust_status": "verified"},
    {"id": "param_total_profit", "symbol": "P", "name": "maximum total profit", "value": total_profit, "unit": "currency", "source_type": "computed_from_data", "source_ref": source_ref, "trust_status": "verified"},
]
result = {
    "id": "result_product_mix_optimum",
    "result_type": "optimization",
    "claim_text": "The selected product quantities maximize fixture profit under all listed constraints.",
    "metrics": {
        "objective_value": objective_value,
        "total_profit": total_profit,
        "material_a_usage": float(usage[0]),
        "material_b_usage": float(usage[1]),
    },
    "source_data": [source_ref],
}
summary = {
    "verified_count": 1,
    "blocked_count": 0,
    "coverage_status": "verified",
    "reader_decision_guide": [
        "Observed data are product profits, resource usage, capacities, and demand limits.",
        "Unknowns are product quantities; the linear program directly identifies the optimal mix under listed constraints.",
        "Time context is not applicable because this fixture is a static product-mix decision.",
    ],
    "method_route_comparison": [
        {
            "route": "linear programming",
            "decision": "recommended",
            "reason": "objective and constraints are linear",
            "boundary": "valid for the listed deterministic capacities and profits",
            "evidence": [result["id"]],
        },
        {
            "route": "nonlinear heuristic search",
            "decision": "rejected",
            "reason": "the objective and constraints are already linear and exactly solvable",
            "boundary": "unnecessary unless the problem introduces nonlinear costs or uncertainty",
            "evidence": [result["id"]],
        }
    ],
    "identifiability_analysis": {
        "observed": "profit coefficients and constraint matrix",
        "unknowns": "product quantities",
        "rank_check": "constraint set yields a bounded optimum",
        "conclusion": "optimization answer is determined by the registered fixture data",
    },
    "claim_registry": [
        {
            "claim": result["claim_text"],
            "status": "evidence_verified",
            "evidence": [result["id"]],
        }
    ],
    "final_answer_tables": [
        {
            "table": "optimal product mix",
            "status": "evidence_verified",
            "result_id": result["id"],
        }
    ],
}

constraint_values = [
    {"name": "material_a_capacity", "value": float(usage[0]), "upper_bound": float(capacity[0])},
    {"name": "material_b_capacity", "value": float(usage[1]), "upper_bound": float(capacity[1])},
    {"name": "product_a_demand", "value": float(best[0]), "upper_bound": float(demand[0])},
    {"name": "product_b_demand", "value": float(best[1]), "upper_bound": float(demand[1])},
    {"name": "product_a_nonnegative", "value": float(best[0]), "lower_bound": 0.0},
    {"name": "product_b_nonnegative", "value": float(best[1]), "lower_bound": 0.0},
]
solver_results = {"evidence": [{
    "kind": "optimization",
    "subproblem_id": "problem_1",
    "result_id": result["id"],
    "objective_value": objective_value,
    "baseline_objective": baseline_objective,
    "constraints": constraint_values,
    "sensitivity": {
        "method": "5 percent capacity stress check",
        "max_relative_change": 0.075,
        "threshold": 0.20,
    },
}]}

Path("parameter_registry.json").write_text(json.dumps({"parameters": parameters}, ensure_ascii=False, indent=2), encoding="utf-8")
Path("result_registry.json").write_text(json.dumps({"verified_results": [result], "blocked_results": [], "summary": summary}, ensure_ascii=False, indent=2), encoding="utf-8")
Path("solver_results.json").write_text(json.dumps(solver_results, ensure_ascii=False, indent=2), encoding="utf-8")
pd.DataFrame(parameters).to_csv("parameter_table.csv", index=False)

Path("figures").mkdir(exist_ok=True)
x_values = np.linspace(0.0, 45.0, 100)
fig, ax = plt.subplots(figsize=(8, 5))
ax.plot(x_values, capacity[0] - 2.0 * x_values, label="material A boundary")
ax.plot(x_values, (capacity[1] - x_values) / 3.0, label="material B boundary")
ax.axvline(demand[0], linestyle="--", label="product A demand")
ax.axhline(demand[1], linestyle="--", label="product B demand")
ax.scatter([best[0]], [best[1]], color="red", zorder=3, label="optimum")
ax.set_xlim(0, 45)
ax.set_ylim(0, 55)
ax.set_xlabel("Product A quantity")
ax.set_ylabel("Product B quantity")
ax.legend()
fig.tight_layout()
fig.savefig("figures/product_mix_feasible_region.png", dpi=160)
plt.close(fig)
figure_registry = {"figures": [{"path": "figures/product_mix_feasible_region.png", "role": "verification_diagnostic", "source_data": [source_ref], "linked_result_ids": [result["id"]]}]}
Path("figure_registry.json").write_text(json.dumps(figure_registry, ensure_ascii=False, indent=2), encoding="utf-8")
print("optimization solver finished")
'''.strip()


def _write_product_mix_xlsx(path: Path) -> None:
    workbook = Workbook()
    products = workbook.active
    products.title = "products"
    products.append(["product", "unit_profit", "material_a", "material_b", "max_demand"])
    products.append(["A", 30, 2, 1, 40])
    products.append(["B", 40, 1, 3, 50])
    capacities = workbook.create_sheet("capacities")
    capacities.append(["resource", "capacity"])
    capacities.append(["material_a", 100])
    capacities.append(["material_b", 90])
    workbook.save(path)


def test_optimization_fixture_runs_same_six_stage_guidance_without_wuyi_specific_text(tmp_path: Path) -> None:
    from app.guidance.execution import LocalProgramExecutor
    from app.guidance.pipeline import GuidancePipeline
    from app.guidance.stages import (
        CalculationStage,
        DecompositionStage,
        GuidanceStage,
        ModelingStage,
        ModelReviewStage,
        ResultVerificationStage,
    )

    data_dir = tmp_path / "data"
    data_dir.mkdir()
    _write_product_mix_xlsx(data_dir / "optimization.xlsx")
    pipeline = GuidancePipeline(
        decomposition_stage=DecompositionStage(decomposer=ProductMixOptimizationDecomposer()),
        modeling_stage=ModelingStage(modeler=ProductMixOptimizationModeler()),
        model_review_stage=ModelReviewStage(reviewer=ProductMixOptimizationReviewer()),
        calculation_stage=CalculationStage(
            program_generator=ProductMixOptimizationProgramGenerator(),
            executor=LocalProgramExecutor(),
        ),
        result_verification_stage=ResultVerificationStage(),
        guidance_stage=GuidanceStage(),
    )
    task = {
        "task_id": "product-mix-optimization",
        "question": "Use the attached workbook to recommend a profit-maximizing product mix.",
        "work_dir": str(tmp_path),
    }

    async def collect():
        return [message async for message in pipeline.run("product-mix-optimization", task)]

    messages = asyncio.run(collect())

    parameters = json.loads((tmp_path / "parameter_registry.json").read_text(encoding="utf-8"))
    parameter_values = {item["id"]: item["value"] for item in parameters["parameters"]}
    results = json.loads((tmp_path / "result_registry.json").read_text(encoding="utf-8"))
    verification = json.loads((tmp_path / "verification_report.json").read_text(encoding="utf-8"))
    guidance = (tmp_path / "guidance.md").read_text(encoding="utf-8")
    audit = json.loads((tmp_path / "guidance_audit_report.json").read_text(encoding="utf-8"))

    assert abs(parameter_values["param_product_a_quantity"] - 40.0) < 1e-8
    assert abs(parameter_values["param_product_b_quantity"] - (50.0 / 3.0)) < 1e-8
    assert abs(parameter_values["param_total_profit"] - (5600.0 / 3.0)) < 1e-8
    assert results["verified_results"][0]["metrics"]["objective_value"] < 0
    assert results["verified_results"][0]["metrics"]["total_profit"] > 0
    assert verification["status"] == "verified"
    assert verification["constraint_checks"][0]["status"] == "passed"
    assert verification["sensitivity_checks"][0]["status"] == "passed"
    assert (tmp_path / "execution_manifest.json").is_file()
    assert json.loads((tmp_path / "execution_manifest.json").read_text(encoding="utf-8"))["execution_count"] == 1
    assert (tmp_path / "notebook.ipynb").is_file()
    assert (tmp_path / "figures" / "product_mix_feasible_region.png").stat().st_size > 1000
    assert "linear programming" in guidance
    assert "result_product_mix_optimum" in guidance
    assert "param_total_profit" in guidance
    assert "data/optimization.xlsx" in guidance
    assert "支路分解" not in guidance
    assert "识别约束" not in guidance
    assert audit["status"] == "PASS"
    assert messages[-1]["data"]["status"] == "completed"


class DemandForecastDecomposer:
    async def decompose(self, *, question: str, data_files: list[str], task: dict):
        from app.guidance.contracts import ProblemSpec

        return ProblemSpec(
            task_summary=question,
            subproblems=[{
                "id": "problem_1",
                "objective": "Forecast next-period demand from the uploaded time-series workbook.",
                "problem_type": "time-series forecasting",
            }],
            data_sources=[{"path": data_files[0], "table": "demand"}],
            locked_facts=[
                "Demand is observed by period in the workbook.",
                "The fixture reserves the final observed period as a holdout check.",
            ],
            uncertainties=["The linear-trend fixture is a first-pass baseline, not a seasonal model."],
        )


class DemandForecastModeler:
    async def model(self, *, problem_spec, task: dict, work_dir: Path):
        from app.guidance.contracts import ModelSpec

        return ModelSpec(
            model_id="demand-forecast-fixture",
            selected_method="linear trend forecast with holdout validation",
            candidate_methods=[
                {
                    "name": "linear trend forecast",
                    "selected": True,
                    "reason": "the fixture data follows a monotone trend and exposes a holdout period",
                },
                {
                    "name": "moving average",
                    "selected": False,
                    "reason": "moving averages lag deterministic trend data",
                },
            ],
            assumptions=[
                "The short-horizon demand trend remains linear for one forecast period.",
                "The final observed period is not used for fitting and is used as a holdout check.",
            ],
            symbols=[
                {"symbol": "t", "meaning": "period index"},
                {"symbol": "D_t", "meaning": "observed demand at period t"},
                {"symbol": "F_9", "meaning": "forecast demand for period 9"},
            ],
            subproblem_models=[{
                "subproblem_id": "problem_1",
                "objective": "Estimate a trend and forecast one period beyond the uploaded data.",
                "selected_method": "linear trend forecast",
                "rationale": "A holdout period verifies the trend equation before the next-period forecast is trusted.",
                "equations": [
                    "D_t = beta_0 + beta_1 t + epsilon_t",
                    "F_9 = beta_0 + beta_1 * 9",
                ],
                "parameters": [
                    {
                        "parameter_id": "trend_intercept",
                        "symbol": "beta_0",
                        "meaning": "linear trend intercept",
                        "unit": "demand units",
                        "source_type": "estimated",
                        "source_detail": "data/forecast.xlsx demand sheet",
                        "estimation_method": "least squares on periods 1 through 7",
                        "constraints": ["finite"],
                    },
                    {
                        "parameter_id": "trend_slope",
                        "symbol": "beta_1",
                        "meaning": "linear trend slope",
                        "unit": "demand units/period",
                        "source_type": "estimated",
                        "source_detail": "data/forecast.xlsx demand sheet",
                        "estimation_method": "least squares on periods 1 through 7",
                        "constraints": ["beta_1 > 0"],
                    },
                ],
                "constraints": ["holdout absolute error <= 1e-8", "forecast demand >= 0"],
                "solve_steps": [
                    "Read the demand sheet from the XLSX attachment.",
                    "Fit a linear trend on periods 1 through 7.",
                    "Validate the fitted trend against period 8.",
                    "Register the next-period forecast and forecast verification metrics.",
                ],
                "expected_results": ["next-period forecast", "holdout RMSE", "trend coefficients"],
            }],
            figure_requests=[{
                "path": "figures/demand_forecast_fit.png",
                "role": "result_fit",
                "linked_result_ids": ["result_demand_forecast"],
                "source_data": ["data/forecast.xlsx"],
            }],
        )


class DemandForecastReviewer:
    async def review(self, *, problem_spec, model_spec, task: dict, work_dir: Path):
        from app.guidance.contracts import ModelReview

        return ModelReview(
            model_id=model_spec.model_id,
            status="approved",
            identifiability="passed_for_linear_trend_baseline",
            dimensional_consistency="passed",
            data_sufficiency="passed_with_holdout_period",
            computational_feasibility="passed",
        )


class DemandForecastProgramGenerator:
    async def generate(self, *, approved_model, task: dict, work_dir: Path, **kwargs):
        from app.guidance.contracts import GeneratedProgram

        return GeneratedProgram(
            source_code=DEMAND_FORECAST_SOLVER,
            declared_outputs=approved_model.required_outputs,
        )


DEMAND_FORECAST_SOLVER = r'''
from pathlib import Path
import json
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt

source = Path("data/forecast.xlsx")
source_ref = source.as_posix()
frame = pd.read_excel(source, sheet_name="demand")
period = frame["period"].astype(float).to_numpy()
demand = frame["demand"].astype(float).to_numpy()

train_period = period[:-1]
train_demand = demand[:-1]
holdout_period = period[-1:]
holdout_observed = demand[-1:]
design = np.column_stack([np.ones_like(train_period), train_period])
coefficients, _, _, _ = np.linalg.lstsq(design, train_demand, rcond=None)
intercept = float(coefficients[0])
slope = float(coefficients[1])
train_predicted = design @ coefficients
holdout_predicted = np.column_stack([np.ones_like(holdout_period), holdout_period]) @ coefficients
next_period = float(period[-1] + 1.0)
forecast_value = float(intercept + slope * next_period)
train_residuals = train_demand - train_predicted
holdout_residuals = holdout_observed - holdout_predicted
train_rmse = float(np.sqrt(np.mean(train_residuals ** 2)))
holdout_rmse = float(np.sqrt(np.mean(holdout_residuals ** 2)))
max_abs_holdout_error = float(np.max(np.abs(holdout_residuals)))
holdout_tolerance = 1e-8

parameters = [
    {"id": "param_trend_intercept", "symbol": "beta_0", "name": "trend intercept", "value": intercept, "unit": "demand units", "source_type": "estimated_from_data", "source_ref": source_ref, "trust_status": "estimated"},
    {"id": "param_trend_slope", "symbol": "beta_1", "name": "trend slope", "value": slope, "unit": "demand units/period", "source_type": "estimated_from_data", "source_ref": source_ref, "trust_status": "estimated"},
    {"id": "param_next_period", "symbol": "t_next", "name": "forecast target period", "value": next_period, "unit": "period", "source_type": "computed_from_data", "source_ref": source_ref, "trust_status": "verified"},
    {"id": "param_next_period_forecast", "symbol": "F_9", "name": "next-period demand forecast", "value": forecast_value, "unit": "demand units", "source_type": "computed_from_data", "source_ref": source_ref, "trust_status": "verified"},
    {"id": "param_holdout_tolerance", "symbol": "tau_holdout", "name": "holdout error tolerance", "value": holdout_tolerance, "unit": "demand units", "source_type": "explicit_model_assumption", "source_ref": source_ref, "trust_status": "assumed"},
]
result = {
    "id": "result_demand_forecast",
    "result_type": "forecast",
    "claim_text": "The linear trend predicts the next-period fixture demand after passing the holdout check.",
    "metrics": {
        "forecast_value": forecast_value,
        "holdout_rmse": holdout_rmse,
        "train_rmse": train_rmse,
    },
    "source_data": [source_ref],
}
summary = {
    "verified_count": 1,
    "blocked_count": 0,
    "coverage_status": "verified",
    "reader_decision_guide": [
        "Observed data are historical demand by period; the unknown is next-period demand.",
        "Time context is the period index in the worksheet.",
        "The forecast is a trend extrapolation supported by holdout error, not a guaranteed future observation.",
    ],
    "method_route_comparison": [
        {
            "route": "linear trend with holdout verification",
            "decision": "recommended",
            "reason": "fixture data follow a linear trend and holdout error is disclosed",
            "boundary": "valid only while the trend assumption remains plausible",
            "evidence": [result["id"]],
        },
        {
            "route": "black-box forecast without holdout check",
            "decision": "rejected",
            "reason": "it would not explain the trend assumption or disclose validation error",
            "boundary": "only useful as an informal baseline",
            "evidence": [result["id"]],
        }
    ],
    "identifiability_analysis": {
        "observed": "period-demand pairs",
        "unknowns": "trend intercept, trend slope, next demand",
        "rank_check": "linear design matrix has two independent columns",
        "conclusion": "trend parameters are identifiable from the fixture series",
    },
    "claim_registry": [
        {
            "claim": result["claim_text"],
            "status": "evidence_verified",
            "evidence": [result["id"]],
        }
    ],
    "final_answer_tables": [
        {
            "table": "next-period forecast",
            "status": "evidence_verified",
            "result_id": result["id"],
        }
    ],
}

solver_results = {"evidence": [{
    "kind": "forecast",
    "subproblem_id": "problem_1",
    "result_id": result["id"],
    "train_observed": train_demand.tolist(),
    "train_predicted": train_predicted.tolist(),
    "holdout_observed": holdout_observed.tolist(),
    "holdout_predicted": holdout_predicted.tolist(),
    "forecast_periods": [next_period],
    "forecast_values": [forecast_value],
    "holdout_tolerance": holdout_tolerance,
    "stability": {
        "method": "endpoint leave-one-out trend refit",
        "max_relative_change": 0.0,
        "threshold": 0.05,
    },
}]}

Path("parameter_registry.json").write_text(json.dumps({"parameters": parameters}, ensure_ascii=False, indent=2), encoding="utf-8")
Path("result_registry.json").write_text(json.dumps({"verified_results": [result], "blocked_results": [], "summary": summary}, ensure_ascii=False, indent=2), encoding="utf-8")
Path("solver_results.json").write_text(json.dumps(solver_results, ensure_ascii=False, indent=2), encoding="utf-8")
pd.DataFrame(parameters).to_csv("parameter_table.csv", index=False)

Path("figures").mkdir(exist_ok=True)
fig, ax = plt.subplots(figsize=(8, 5))
ax.plot(period, demand, "o", label="observed demand")
ax.plot(train_period, train_predicted, "-", label="fitted trend")
ax.scatter(holdout_period, holdout_predicted, marker="x", s=80, label="holdout prediction")
ax.scatter([next_period], [forecast_value], color="red", label="next forecast")
ax.set_xlabel("Period")
ax.set_ylabel("Demand")
ax.legend()
fig.tight_layout()
fig.savefig("figures/demand_forecast_fit.png", dpi=160)
plt.close(fig)
figure_registry = {"figures": [{"path": "figures/demand_forecast_fit.png", "role": "result_fit", "source_data": [source_ref], "linked_result_ids": [result["id"]]}]}
Path("figure_registry.json").write_text(json.dumps(figure_registry, ensure_ascii=False, indent=2), encoding="utf-8")
print("forecast solver finished")
'''.strip()


def _write_forecast_xlsx(path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "demand"
    sheet.append(["period", "demand"])
    for period in range(1, 9):
        sheet.append([period, 100 + 12 * period])
    workbook.save(path)


def test_forecast_fixture_runs_same_six_stage_guidance_with_holdout_verification(tmp_path: Path) -> None:
    from app.guidance.execution import LocalProgramExecutor
    from app.guidance.pipeline import GuidancePipeline
    from app.guidance.stages import (
        CalculationStage,
        DecompositionStage,
        GuidanceStage,
        ModelingStage,
        ModelReviewStage,
        ResultVerificationStage,
    )

    data_dir = tmp_path / "data"
    data_dir.mkdir()
    _write_forecast_xlsx(data_dir / "forecast.xlsx")
    pipeline = GuidancePipeline(
        decomposition_stage=DecompositionStage(decomposer=DemandForecastDecomposer()),
        modeling_stage=ModelingStage(modeler=DemandForecastModeler()),
        model_review_stage=ModelReviewStage(reviewer=DemandForecastReviewer()),
        calculation_stage=CalculationStage(
            program_generator=DemandForecastProgramGenerator(),
            executor=LocalProgramExecutor(),
        ),
        result_verification_stage=ResultVerificationStage(),
        guidance_stage=GuidanceStage(),
    )
    task = {
        "task_id": "demand-forecast",
        "question": "Use the attached workbook to forecast next-period demand.",
        "work_dir": str(tmp_path),
    }

    async def collect():
        return [message async for message in pipeline.run("demand-forecast", task)]

    messages = asyncio.run(collect())

    parameters = json.loads((tmp_path / "parameter_registry.json").read_text(encoding="utf-8"))
    parameter_values = {item["id"]: item["value"] for item in parameters["parameters"]}
    results = json.loads((tmp_path / "result_registry.json").read_text(encoding="utf-8"))
    verification = json.loads((tmp_path / "verification_report.json").read_text(encoding="utf-8"))
    guidance = (tmp_path / "guidance.md").read_text(encoding="utf-8")
    audit = json.loads((tmp_path / "guidance_audit_report.json").read_text(encoding="utf-8"))

    assert abs(parameter_values["param_trend_intercept"] - 100.0) < 1e-8
    assert abs(parameter_values["param_trend_slope"] - 12.0) < 1e-8
    assert parameter_values["param_next_period"] == 9.0
    assert abs(parameter_values["param_next_period_forecast"] - 208.0) < 1e-8
    assert abs(results["verified_results"][0]["metrics"]["forecast_value"] - 208.0) < 1e-8
    assert results["verified_results"][0]["metrics"]["holdout_rmse"] < 1e-8
    assert verification["status"] == "verified"
    assert verification["residual_checks"][0]["status"] == "passed"
    assert verification["sensitivity_checks"][0]["status"] == "passed"
    assert (tmp_path / "execution_manifest.json").is_file()
    assert json.loads((tmp_path / "execution_manifest.json").read_text(encoding="utf-8"))["execution_count"] == 1
    assert (tmp_path / "notebook.ipynb").is_file()
    assert (tmp_path / "figures" / "demand_forecast_fit.png").stat().st_size > 1000
    assert "linear trend forecast" in guidance
    assert "result_demand_forecast" in guidance
    assert "param_next_period_forecast" in guidance
    assert "data/forecast.xlsx" in guidance
    assert audit["status"] == "PASS"
    assert messages[-1]["data"]["status"] == "completed"
